"""
Shorts Check - Streamlit 대시보드
기획서 8-3항 구현: Verdict Card, Reasoning View, Metric Gauge
"""

import streamlit as st
from PIL import Image
import io as _io
import base64 as _b64
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import requests
import random
from datetime import datetime, timedelta
import logging
import os

# ─────────────────────────────────────────
# 페이지 설정 (반드시 최상단)
# ─────────────────────────────────────────
# 로고 base64
_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAACAAAAAgCAYAAABzenr0AAAIgklEQVR4nK2XbYycVRXHf/feZ152d2Zndru7bbcv21dot9BtCylCSho0Rj9AkEg/AAVBMFHBthChpoAaRUmICgjGYIBAIopU+AQlpcVSqLSEYjGkFYiA9L273e2+TGf2mbn3HD/Ms7stASMvN7l5Zj7c+/+fc//nf881fIFDwbBypTPbt3uA7u7u3L253PKs9zOlZkovNEav3P3aa8cUrAGtL/mCgLetXBmN/f/unDkd+xYv/3HvgiXvnZh5ho5Mnq3DHbP1aNfCg9sWLb16nOwXME4DXtvVVXyjZ+kdRxec3atT5+tIuk2P0SxHaA6HyId+U9Bq2yz917yFjwGRrlrlPjPwqYsv7eoq7uk5Z/3xM3r2a+c8HU636mFytUMmL0dsQY+4Fj3kWnS/LcpBclWdMlf/0X3WXfDp02B01SprN24MClww6cz87zrz13cGv7ZjuDxruHeAcrXixThnjTGioCgYO75BUJV8lNbqrI7S46O++/8lcBowXV3Z3cVJ18+oyi0dw6Nzysf7GY5HvWCcscao6ri6FDAuwliLhICqIIJ2TGs3bzRG34w+GfN0YLNxYwAye5Yuv7bTV9d2DJQWlvuH6C2PegGLNREKooABVVAMxlqCH0IJGFIoDQgqNoj1QWZ+EoHTgc8htcP3XDs3Zl1H/3B33NfPsUolKM4Y6yLUI2om6koBDOossT9B0xXXkVl6NqUnnkTe/TehEmMUUxGaP3oEZtvKle7L27f7JIWZnWede+Ws4Ne1nawsjvsGGaqUg8EYa51VQFURBMWQnHj9tzPEfpDculuZfO89WCCuVvlg6Xmk9u0NM6bOcC9l3J3jGVCwFuSiuomkXltyzqpp1bBhylBpUXx8gL5KOSjGGGucqiEkp6yGJHoFBbEGNUrshyje8TMm//xOZDRGsxne+eEGUvv+CTYHIWBxRABPgTMQgOjFhYuvObPGmo7jIz1hYJi+ciUIYqx1ThREZUJdSbSSRI61CJ7RUKblnnuZcus6auUyqcZG3nn4MeIHfk0+asb7gISA+ECU2GJ4Zu5Z5y+P7INThuNl1YFB+pOIrbXOGocaA6Ko2gRYUQw6Bm8dGmLKNtDx6OO0XbeauFQik8tx4NVdHLnpJrpcDq8ggKrBApEBebpr3pe+EsLWXO9AY+/IiA9grbXOKKi1iK8QGMWQBZdFNTBW44qizhF8hVpDhql/+gst37iE6kiJqCHLicNH2HvFambGVWquARVBsXXSIthV06c3LAvmkVzvQOOxkeGatTYyxlhRIRio+gH8tBnk//AwZmE3IQwiUQqhLr7gHDVfYrS1QOfzz9bBSyVsKiKosuvKa2jZ/x7pqAkvvh49iojUCVxL+rLOSrW7r1wOxrqUT0xErMFrBXfxpXTsfJHm73yb1mc3IgvOplY9gUQRwaWI/TDVGdOZvfUF8isvJB4poYDLZtl5y21Ut2+lkCpQDVUEiwDegKgiXrCTa7KCyqgGICgIhqAQjKOmVcLiZWRmTKc6OISbM5uOLZtgyTLi2iCxH8IsXMSsF7eQXdpDdWQEUSGTy/HWQ49w8MH76EzlCD7gqVdODcVTJwBg81XJiA9GkpJSA2IgBI+1OY7efReHtr9MqligNjhM1DmVzs2b0CXL0J6lzP7bFjJz5xAPDRFUSeXzfPjKDt68eS1TXJaMQEAJJOJLZlCh5j1Rn9QOTzcgSV1JomsBsBE5X+GDm9fT+uo2bDpFKJUxLS10Pf8sqsCkVvxICYwllUlTOnqU7Vd/i1ylQt41gQhhHDg5XhLLBuzb1J4ZdFYjY6ipakjAFYOEQCpqxuzZxbu/uZ+osQERj1QqmEIBW2gmlMt180ERY9h8zXXoh+9TTDUSScCjhCSosb0DiorgAHvD0PE97zvzUGcm66xKTRTR8URBkEDBNnLgl3fTv3cfUVMTooLWaqj3SZl6ss3N7Fh/Oye2bqY1lSPrAwLUAJ/MgCEkRyyiBAGrYC8cObb2DadPT01n0jmDVTQoKgZFVbA2Rf7kEG/ffBtqE0NKKPo4pqGlhT2PPMq++35Fe6qJbPBEmAS0HrFPpmj9G7ReijbRXfXck8cv35Gya0pR6j/tNu0KYEU1BBUheFqjZuItz3H4j0+SaSkSRmN8HJNtKfLhKzvY8YM1FFyaSIRGNeMpHxPgBBGhpkJAsYClfnUbBfPVk30PnJ83i1+P7I0DUbSv1UWu1RiLqljVULAZ9v9oA8PvvEuuo53m9nZKBw+zefXVZCpl0jgatL6ph3ECYiaE7U+pBuEjLZlOXErQTXrT+8UrOoU1bUGWWVWOgQxLWcvTZ7rc7etJF4vsvusXxHvfouCyFERpMxZ0Qu0TnVFSWfVvmB81ujdT0Z0f15KZ5GoOyWKzsaFw2azg1k0K4UIhcCSUdS8iB8C2gsm5NBlRpmBpMOB1ot5PnWMEAoR5UaPb7ewd/6snNPU7jpBcwDyRaflal4SbCtiL0ygv6ygjKqFJ1BaNNe1qTjObCV+pO+xYA+MNMj9qMPujzPftxyGPZc0k4E+BUzBXxSc2r6gNX7LDmRV9xv35TFK+W61rApNTfEBVdKLm/alT6wL0pm5HAYwPvv9TteVPgVuVuDXA7xuLPQu8fi8nelVeQq5XaowgXoxxgKlHfYr7AQGjzRiZHDXGWxwLP9PzKBG6GRPs/c3tcxfF1TWFEFZnJbT2aWCwLgVrqHcwiqEGKiq1C1xT+u0o+u3X4/61n+t99hOwPz2FyIaGhmkrJH1jUbghJ769XzyDCAEVUUPGYOdHDfTZ1Kar2vsvP3CQ+PPgn0ZEYfypdlnT5I7n0i23/T3K793l8vq6y+nuKK87U4Vjz2XbboHx+jdfyAt1bCiYl8BdVPcbmEfmrwfbzsuG2hwid3Iz7HygMnBQJ4D1v+NAz0sGCK8tAAAAAElFTkSuQmCC"

# ── 파비콘: PIL Image 방식 (Streamlit 공식 지원) ──
import base64 as _b64c, io as _ioc
from PIL import Image as _PILI
_fav_img = _PILI.open(_ioc.BytesIO(_b64c.b64decode(_LOGO_B64))).convert("RGBA")

st.set_page_config(
    page_title="Shorts Check",
    page_icon="🔴",
    layout="wide",
    initial_sidebar_state="collapsed"
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def _inject_favicon(b64: str):
    """JS로 브라우저 파비콘을 직접 교체 - 가장 확실한 방법"""
    st.components.v1.html(f"""
    <script>
    (function() {{
      var b64 = "{b64}";
      var url = "data:image/png;base64," + b64;
      
      // 기존 favicon 제거
      var links = document.querySelectorAll("link[rel*='icon']");
      links.forEach(function(l) {{ l.parentNode.removeChild(l); }});
      
      // 새 favicon 삽입
      var link = document.createElement('link');
      link.type = 'image/png';
      link.rel = 'shortcut icon';
      link.href = url;
      document.getElementsByTagName('head')[0].appendChild(link);
      
      // 탭 아이콘 즉시 갱신
      var link2 = document.createElement('link');
      link2.type = 'image/png';  
      link2.rel = 'icon';
      link2.href = url;
      document.getElementsByTagName('head')[0].appendChild(link2);
    }})();
    </script>
    """, height=0, scrolling=False)


# ─────────────────────────────────────────
# 전역 CSS 스타일
# ─────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');

  html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
  }

  /* ── 4. 전체 배경 흰색 ── */
  .stApp {
    background: #ffffff !important;
    color: #1a1a2e;
  }

  /* ── 상단 헤더/툴바 흰색 ── */
  [data-testid="stToolbar"],
  header[data-testid="stHeader"],
  .stDeployButton {
    background: #ffffff !important;
  }
  header[data-testid="stHeader"] {
    background: #ffffff !important;
    border-bottom: 1px solid rgba(0,0,0,0.08) !important;
  }
  [data-testid="stToolbar"] button,
  [data-testid="stToolbar"] a,
  header[data-testid="stHeader"] button,
  header[data-testid="stHeader"] a,
  .stDeployButton button {
    color: #888888 !important;
    background: transparent !important;
    transition: color 0.2s ease !important;
  }
  [data-testid="stToolbar"] button:hover,
  header[data-testid="stHeader"] button:hover,
  .stDeployButton button:hover {
    color: #333333 !important;
    background: rgba(0,0,0,0.05) !important;
  }

  /* ── 2. 헤더 타이틀 가운데 정렬 ── */
  .main-title {
    font-family: 'Syne', sans-serif;
    font-size: 2.6rem;
    font-weight: 800;
    letter-spacing: -1px;
    color: #1a1a2e;
    margin-bottom: 0;
    line-height: 1.1;
    text-align: center;
  }
  .main-title .title-red {
    color: #cc0000;
  }
  .main-subtitle {
    font-family: 'DM Sans', sans-serif;
    font-size: 0.95rem;
    color: #888888;
    margin-top: 4px;
    letter-spacing: 0.02em;
    text-align: center;
  }
  .main-header-wrap {
    text-align: center;
    margin-bottom: 24px;
  }

  /* ── Verdict 카드 ── */
  .verdict-card {
    border-radius: 16px;
    padding: 28px 32px;
    position: relative;
    overflow: hidden;
    margin-bottom: 8px;
  }
  .verdict-danger  { background: linear-gradient(135deg, rgba(239,68,68,0.10), rgba(185,28,28,0.06)); border-left: 4px solid #ef4444; }
  .verdict-warning { background: linear-gradient(135deg, rgba(245,158,11,0.10), rgba(180,83,9,0.06));  border-left: 4px solid #f59e0b; }
  .verdict-safe    { background: linear-gradient(135deg, rgba(16,185,129,0.10), rgba(5,150,105,0.06));  border-left: 4px solid #10b981; }

  .verdict-category {
    font-family: 'Syne', sans-serif;
    font-size: 3.2rem;
    font-weight: 800;
    line-height: 1;
    margin: 0;
  }
  .verdict-label {
    font-family: 'DM Sans', sans-serif;
    font-size: 1.1rem;
    font-weight: 500;
    margin: 6px 0 0 0;
    opacity: 0.85;
  }
  .verdict-badge {
    display: inline-block;
    font-size: 0.72rem;
    font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    padding: 3px 10px;
    border-radius: 20px;
    margin-top: 10px;
  }
  .badge-danger  { background: rgba(239,68,68,0.15);  color: #dc2626; }
  .badge-warning { background: rgba(245,158,11,0.15); color: #d97706; }
  .badge-safe    { background: rgba(16,185,129,0.15); color: #059669; }

  /* ── Metric 카드 ── */
  .metric-card {
    background: #f8f9fa;
    border: 1px solid #e9ecef;
    border-radius: 14px;
    padding: 20px 22px;
    text-align: center;
  }
  .metric-value {
    font-family: 'Syne', sans-serif;
    font-size: 2.1rem;
    font-weight: 700;
    line-height: 1;
    color: #333333;
  }
  .metric-label {
    font-size: 0.78rem;
    color: #888888;
    margin-top: 6px;
    letter-spacing: 0.05em;
    text-transform: uppercase;
  }
  .metric-weight {
    font-size: 0.7rem;
    color: #aaaaaa;
    margin-top: 3px;
  }

  /* ── Reasoning 박스 ── */
  .reasoning-box {
    background: #fff8f8;
    border: 1px solid rgba(204,0,0,0.15);
    border-radius: 12px;
    padding: 18px 22px;
    font-size: 0.92rem;
    line-height: 1.75;
    color: #444444;
  }

  /* ── 섹션 헤더 ── */
  .section-header {
    font-family: 'Syne', sans-serif;
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 0.14em;
    text-transform: uppercase;
    color: #aaaaaa;
    margin: 28px 0 12px 0;
  }

  /* ── 3. 버튼: 빨간색 계열 ── */
  .stButton > button {
    background: #ffffff !important;
    border: 1.5px solid #cc0000 !important;
    color: #cc0000 !important;
    border-radius: 10px !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.88rem !important;
    font-weight: 600 !important;
    padding: 10px 16px !important;
    transition: all 0.2s ease !important;
    width: 100%;
  }
  .stButton > button:hover,
  .stButton > button:active,
  .stButton > button:focus {
    background: #cc0000 !important;
    border-color: #cc0000 !important;
    color: #ffffff !important;
    transform: translateY(-1px);
  }

  /* ── 3. 분석 submit 버튼도 빨간색 ── */
  [data-testid="stFormSubmitButton"] > button {
    background: #cc0000 !important;
    border: 1.5px solid #cc0000 !important;
    color: #ffffff !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    transition: all 0.2s ease !important;
  }
  [data-testid="stFormSubmitButton"] > button:hover {
    background: #aa0000 !important;
    border-color: #aa0000 !important;
    color: #ffffff !important;
    transform: translateY(-1px);
  }

  /* ── 1. 사이드바 토글 가능하게 (collapsed 상태 스타일) ── */
  [data-testid="stSidebar"] {
    background: #f8f9fa !important;
    border-right: 1px solid #e9ecef !important;
    transition: all 0.3s ease !important;
  }
  [data-testid="stSidebar"][aria-expanded="false"] {
    display: none !important;
  }
  /* 사이드바 토글 버튼 스타일 */
  [data-testid="collapsedControl"] {
    color: #cc0000 !important;
  }
  button[kind="header"] {
    color: #cc0000 !important;
  }

  /* 스크롤바 */
  ::-webkit-scrollbar { width: 6px; height: 6px; }
  ::-webkit-scrollbar-track { background: #f0f0f0; }
  ::-webkit-scrollbar-thumb { background: #999999 !important; border-radius: 4px; }
  ::-webkit-scrollbar-thumb:hover { background: #777777 !important; }

  /* ── 입력창 ── */
  .stTextInput > div > div > input {
    background: #ffffff !important;
    border: 1.5px solid #dddddd !important;
    border-radius: 10px !important;
    color: #BC8F8F !important;
    font-family: 'DM Sans', sans-serif !important;
  }
  .stTextInput > div > div > input::placeholder {
    color: #aaaaaa !important;
    opacity: 1;
  }
  .stTextInput > div > div > input:focus {
    border-color: #cc0000 !important;
    box-shadow: 0 0 0 2px rgba(204,0,0,0.12) !important;
    outline: none !important;
  }
  /* Enter 힌트 숨김 */
  .stTextInput small,
  .stTextInput [data-testid="InputInstructions"],
  small[data-testid="InputInstructions"] {
    display: none !important;
    height: 0 !important;
  }

  /* ── 상태 칩 ── */
  .status-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    font-size: 0.8rem;
    font-weight: 600;
    padding: 5px 12px;
    border-radius: 20px;
  }
  .chip-online  { background: rgba(16,185,129,0.1); color: #059669; border: 1px solid rgba(16,185,129,0.2); }
  .chip-offline { background: rgba(239,68,68,0.1);  color: #dc2626; border: 1px solid rgba(239,68,68,0.2); }

  /* ── 분리선 ── */
  hr { border-color: #e9ecef !important; }

  /* ── OCR 텍스트 박스 ── */
  .stTextArea textarea {
    background: #f8f9fa !important;
    border: 1px solid #e9ecef !important;
    color: #666666 !important;
    font-size: 0.85rem !important;
    border-radius: 10px !important;
  }

  /* ── 테이블 ── */
  .stDataFrame { border-radius: 12px !important; overflow: hidden; }

  /* ── Spinner ── */
  .stSpinner > div { border-top-color: #cc0000 !important; }

  /* ── 알림 박스 ── */
  .stAlert { border-radius: 10px !important; }

  /* ── form 포커스 아웃라인 제거 ── */
  [data-testid="stForm"]:focus-within {
    outline: none !important;
    border-color: transparent !important;
  }

</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────
# 헬퍼 함수들
# ─────────────────────────────────────────

CATEGORY_META = {
    "C1": {"label": "어그로 / 스팸",    "verdict": "danger",  "icon": "🚨", "badge": "HIGH RISK"},
    "C2": {"label": "공장형 패턴",      "verdict": "danger",  "icon": "🏭", "badge": "HIGH RISK"},
    "C3": {"label": "품질 불량",        "verdict": "warning", "icon": "⚠️",  "badge": "REVIEW"},
    "C4": {"label": "무단 도용",        "verdict": "danger",  "icon": "⚖️",  "badge": "HIGH RISK"},
    "C5": {"label": "정상 영상",        "verdict": "safe",    "icon": "✅",  "badge": "SAFE"},
}

ACTION_MAP = {
    "C1": [("👍", "좋아요", "like"), ("🚫", "채널 추천 안 함", "block_channel"), ("📢", "신고하기", "report"), ("💬", "의견 보내기", "feedback")],
    "C2": [("👍", "좋아요", "like"), ("🚫", "채널 추천 안 함", "block_channel"), ("📢", "신고하기", "report"), ("💬", "의견 보내기", "feedback")],
    "C3": [("👍", "좋아요", "like"), ("🚫", "채널 추천 안 함", "block_channel"), ("📢", "신고하기", "report"), ("💬", "의견 보내기", "feedback")],
    "C4": [("👍", "좋아요", "like"), ("🚫", "채널 추천 안 함", "block_channel"), ("📢", "신고하기", "report"), ("💬", "의견 보내기", "feedback")],
    "C5": [("👍", "좋아요", "like"), ("🚫", "채널 추천 안 함", "block_channel"), ("📢", "신고하기", "report"), ("💬", "의견 보내기", "feedback")],
}


def get_verdict_class(category: str) -> str:
    return CATEGORY_META.get(category, {}).get("verdict", "warning")


def _extract_video_id(url: str) -> str:
    import re
    for p in [r'youtube\.com/shorts/([a-zA-Z0-9_-]+)',
              r'youtu\.be/([a-zA-Z0-9_-]+)',
              r'youtube\.com/watch\?v=([a-zA-Z0-9_-]+)']:
        m = re.search(p, url)
        if m:
            return m.group(1)
    return f"video_{abs(hash(url)) % 100000}"


def _is_valid_youtube_url(url: str) -> bool:
    """YouTube URL + video ID 11자리 형식까지 검증"""
    import re
    # YouTube video ID는 정확히 11자리 영문·숫자·_·- 조합
    patterns = [
        r'youtube\.com/shorts/([a-zA-Z0-9_-]{11})(?:[/?&#]|$)',
        r'youtu\.be/([a-zA-Z0-9_-]{11})(?:[/?&#]|$)',
        r'youtube\.com/watch\?v=([a-zA-Z0-9_-]{11})(?:[&# ]|$)',
    ]
    return any(re.search(p, url) for p in patterns)


def _check_video_exists(url: str) -> bool:
    """oEmbed API로 영상 실제 존재 여부 확인. 404면 False, 200이면 True."""
    try:
        import urllib.parse, urllib.request
        oembed_url = ("https://www.youtube.com/oembed?url="
                      + urllib.parse.quote(url) + "&format=json")
        req = urllib.request.Request(oembed_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            return resp.status == 200
    except Exception as e:
        # 404 = 영상 없음, 그 외 = 네트워크 오류 → 존재한다고 가정(분석 진행)
        if "404" in str(e) or "HTTP Error 404" in str(e):
            return False
        return True


def _show_url_error(error_type: str = "invalid", url: str = ""):
    """잘못된 URL 또는 분석 오류 시 커스텀 에러 화면 표시"""
    _, c, _ = st.columns([1, 3, 1])
    with c:
        if error_type == "invalid":
            icon, title, desc, tip = (
                "🔗",
                "올바른 YouTube URL이 아니에요",
                "YouTube Shorts, 일반 영상, 단축 URL만 분석할 수 있어요.",
                [
                    "youtube.com/shorts/xxxxxxxxx",
                    "youtube.com/watch?v=xxxxxxxxx",
                    "youtu.be/xxxxxxxxx",
                ]
            )
            tip_html = "".join(
                f"<div style='font-family:monospace;font-size:0.82rem;"
                f"color:#888;padding:3px 0'>✓ {t}</div>" for t in tip
            )
            st.markdown(f"""
            <div style="background:#fff8f8;border:1.5px solid #fca5a5;border-radius:16px;
                        padding:32px 28px;text-align:center;margin-top:8px">
              <div style="font-size:2.4rem;margin-bottom:12px">{icon}</div>
              <div style="font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;
                          color:#cc0000;margin-bottom:8px">{title}</div>
              <div style="font-size:0.88rem;color:#888;margin-bottom:20px">{desc}</div>
              <div style="background:#fff;border:1px solid #f0e0e0;border-radius:10px;
                          padding:14px 18px;text-align:left;margin-bottom:16px">
                <div style="font-size:0.75rem;font-weight:700;color:#cc000088;
                            letter-spacing:0.08em;margin-bottom:8px">지원하는 URL 형식</div>
                {tip_html}
              </div>
              {"<div style='font-size:0.78rem;color:#bbb;word-break:break-all'>입력값: " + url + "</div>" if url else ""}
            </div>
            """, unsafe_allow_html=True)

        elif error_type == "not_found":
            st.markdown(f"""
            <div style="background:#fffbf0;border:1.5px solid #fcd34d;border-radius:16px;
                        padding:32px 28px;text-align:center;margin-top:8px">
              <div style="font-size:2.4rem;margin-bottom:12px">📭</div>
              <div style="font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;
                          color:#b45309;margin-bottom:8px">영상을 찾을 수 없어요</div>
              <div style="font-size:0.88rem;color:#888;margin-bottom:12px">
                삭제되었거나 비공개 처리된 영상일 수 있어요.
              </div>
              {"<div style='font-size:0.78rem;color:#bbb;word-break:break-all'>입력값: " + url + "</div>" if url else ""}
            </div>
            """, unsafe_allow_html=True)

        elif error_type == "timeout":
            # 피드백 버튼 — session_state로 전송 여부 관리
            if "timeout_feedback_sent" not in st.session_state:
                st.session_state.timeout_feedback_sent = False

            st.markdown(f"""
            <div style="background:#fffbf0;border:1.5px solid #fbbf24;border-radius:16px;
                        padding:32px 28px;text-align:center;margin-top:8px">
              <div style="font-size:2.4rem;margin-bottom:12px">⏱️</div>
              <div style="font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;
                          color:#b45309;margin-bottom:8px">검증 시간이 초과되었어요</div>
              <div style="font-size:0.88rem;color:#777;margin-bottom:6px">
                분석이 65초를 초과했습니다. 서버 상태가 일시적으로 불안정할 수 있어요.
              </div>
              <div style="font-size:0.82rem;color:#aaa;margin-bottom:20px">
                관리자에게 알려주시면 빠르게 확인하겠습니다 🙏
              </div>
              <div style="font-family:monospace;font-size:0.76rem;color:#d97706;
                          background:#fef9ec;border:1px solid #fde68a;
                          border-radius:8px;padding:8px 14px;margin-bottom:20px;
                          word-break:break-all">
                {url}
              </div>
            </div>
            """, unsafe_allow_html=True)

            # 피드백 전송 버튼
            _, btn_col, _ = st.columns([1, 3, 1])
            with btn_col:
                if st.session_state.timeout_feedback_sent:
                    st.markdown("""
                    <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:10px;
                                padding:12px;text-align:center;font-size:0.88rem;color:#16a34a">
                      ✅ 관리자에게 전달했습니다. 감사합니다!
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    if st.button("📢 관리자에게 알리기", key="btn_timeout_feedback",
                                 use_container_width=True, type="primary"):
                        sent = submit_feedback(
                            video_id=f"timeout_{abs(hash(url)) % 100000}",
                            action="report",
                            text=f"[검증시간초과] URL: {url}"
                        )
                        st.session_state.timeout_feedback_sent = True
                        st.rerun()

        else:  # general error
            st.markdown(f"""
            <div style="background:#f8f9ff;border:1.5px solid #c7d2fe;border-radius:16px;
                        padding:32px 28px;text-align:center;margin-top:8px">
              <div style="font-size:2.4rem;margin-bottom:12px">⚡</div>
              <div style="font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;
                          color:#4338ca;margin-bottom:8px">분석 중 오류가 발생했어요</div>
              <div style="font-size:0.88rem;color:#888;margin-bottom:12px">
                잠시 후 다시 시도해주세요. 문제가 반복되면 URL을 확인해주세요.
              </div>
              <div style="font-family:monospace;font-size:0.78rem;color:#a5b4fc;
                          background:#eef2ff;border-radius:8px;padding:8px 12px">
                {url}
              </div>
            </div>
            """, unsafe_allow_html=True)


def _analyze_with_openai_direct(video_url: str, video_id: str) -> dict:
    """
    Railway 없이 Streamlit에서 직접 OpenAI API 호출로 분석
    영상 메타데이터 + GPT-4o-mini 텍스트 분석
    """
    import time
    start = time.time()

    try:
        from openai import OpenAI

        api_key = (
            st.secrets.get("OPENAI_API_KEY", None)
            or os.getenv("OPENAI_API_KEY", "")
        )
        if not api_key:
            raise Exception("OPENAI_API_KEY가 설정되지 않았습니다.")

        client = OpenAI(api_key=api_key)

        # 유튜브 메타데이터 수집
        yt_info = _fetch_youtube_info(video_url)
        title = yt_info.get("title", "—")
        view_count = yt_info.get("view_count", 0)
        duration = yt_info.get("duration_sec", 0)

        # GPT-4o-mini 호출 (텍스트만)
        prompt = f"""아래 유튜브 쇼츠 영상을 분석해주세요.

## 영상 메타데이터
- URL: {video_url}
- 제목: {title}
- 조회수: {view_count}회
- 영상 길이: {duration}초

※ 프레임 이미지 없이 메타데이터만으로 분석합니다.

반드시 아래 JSON만 출력하세요:
{{
    "c_category": "C1~C5 중 하나",
    "confidence_score": 0.0~1.0,
    "subtitle_detected": false,
    "subtitle_content": "자막 없음",
    "frame_analysis": "프레임 미제공",
    "reasoning_log": "판단 근거 상세 설명"
}}"""

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 한국 유튜브 쇼츠 콘텐츠를 분석하는 전문가입니다. C1~C5로 분류하고 반드시 JSON만 출력하세요."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=800,
            temperature=0.1
        )

        raw = response.choices[0].message.content
        # JSON 파싱
        import json, re
        json_match = re.search(r'\{.*\}', raw, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group())
        else:
            raise Exception("JSON 파싱 실패")

        processing_time = time.time() - start
        category = result.get("c_category", "C5")
        confidence = float(result.get("confidence_score", 0.5))
        reasoning = result.get("reasoning_log", "")

        status_map = {
            "C1": "AUTO_REJECT", "C2": "AUTO_REJECT",
            "C3": "HUMAN_REVIEW", "C4": "AUTO_REJECT", "C5": "AUTO_APPROVE"
        }

        return {
            "video_id": video_id,
            "analysis_result": {
                "category": category,
                "confidence_score": confidence,
                "reasoning_log": f"[자막] {result.get('subtitle_content','—')}\n[프레임] {result.get('frame_analysis','—')}\n[판단] {reasoning}",
                "status": status_map.get(category, "AUTO_APPROVE"),
            },
            "context_score": {
                "context_score":    round(confidence * 0.9, 3),
                "s_semantic":       round(confidence * 0.95, 3),
                "o_existence":      round(confidence * 0.85, 3),
                "a_sync":           round(confidence * 0.9, 3),
                "layout_score":     0.0,       # 직접 호출 시 프레임 미추출
                "keyframe_count":   0,          # 직접 호출 시 프레임 미추출
                "spam_detected":    False,
                "short_circuit_c4": category in ["C1", "C2", "C3"],
            },
            "processing_time": processing_time,
            "model_used": "gpt-4o-mini (직접 호출)",
            "video_info": {
                "title": title,
                "view_count": view_count,
                "duration": duration,
            }
        }

    except Exception as e:
        raise Exception(f"OpenAI 직접 분석 실패: {str(e)}")


def _fetch_youtube_title(url: str) -> str:
    """YouTube URL에서 영상 제목을 가져옴 (oEmbed API 사용, 실패 시 '—' 반환)"""
    info = _fetch_youtube_info(url)
    return info.get("title", "—")


def _fetch_youtube_info(url: str) -> dict:
    """YouTube URL에서 제목·조회수·길이를 가져옴. 실패 시 빈 dict 반환."""
    if not url or url == "—":
        return {}
    try:
        import re, json as _json

        vid_id = _extract_video_id(url)
        if not vid_id:
            return {}

        title = "—"
        view_count = None
        duration_sec = None

        # YouTube Data API v3 사용
        yt_api_key = st.secrets.get("YOUTUBE_API_KEY", "") or os.getenv("YOUTUBE_API_KEY", "")
        if yt_api_key:
            try:
                api_url = (
                    f"https://www.googleapis.com/youtube/v3/videos"
                    f"?part=snippet,contentDetails,statistics&id={vid_id}&key={yt_api_key}"
                )
                resp = requests.get(api_url, timeout=8)
                data = resp.json()
                items = data.get("items", [])
                if items:
                    item = items[0]
                    title = item.get("snippet", {}).get("title", "—")
                    view_count = int(item.get("statistics", {}).get("viewCount", 0))
                    # ISO 8601 duration (PT1H2M3S) → 초 변환
                    duration_iso = item.get("contentDetails", {}).get("duration", "PT0S")
                    m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', duration_iso)
                    if m:
                        h, mi, s = (int(x or 0) for x in m.groups())
                        duration_sec = h * 3600 + mi * 60 + s
            except Exception:
                pass
        else:
            # API 키 없을 때 oEmbed로 제목만 취득
            try:
                import urllib.parse, urllib.request
                oembed_url = ("https://www.youtube.com/oembed?url="
                              + urllib.parse.quote(url) + "&format=json")
                req = urllib.request.Request(oembed_url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=5) as r:
                    title = _json.loads(r.read().decode()).get("title", "—")
            except Exception:
                pass

        return {"title": title, "view_count": view_count, "duration_sec": duration_sec}
    except Exception:
        return {}


def make_mock_analysis(video_id: str) -> dict:
    """Mock 분석 데이터 (API/DB 없을 때 사용)"""
    cats = ["C1", "C2", "C3", "C4", "C5"]
    cat = random.choice(cats)
    conf = round(random.uniform(0.55, 0.97), 3)
    s = round(random.uniform(0.5, 0.95), 3)
    o = round(random.uniform(0.4, 0.9), 3)
    a = round(random.uniform(0.5, 0.92), 3)
    ctx = round(s * 0.5 + o * 0.3 + a * 0.2, 3)
    reasoning_map = {
        "C1": "제목에 자극적 키워드('충격', '100만원', '🔥')가 다수 감지되었습니다. 영상 내용과 제목 사이의 의미적 유사도가 낮아 낚시성 콘텐츠로 판단됩니다.",
        "C2": "TTS 특유의 일정한 피치·톤 반복성이 감지되었으며 레이아웃 수치(ROI)가 정적 템플릿 패턴과 일치합니다. 공장형 대량 생산 콘텐츠로 분류합니다.",
        "C3": "OCR 추출 자막과 영상 화면 간 의미적 불일치(S_semantic < 0.6)가 확인되었습니다. 자막-화면 싱크 오류가 품질 문제를 시사합니다.",
        "C4": "역이미지 검색 결과 최초 업로드 시점이 다른 채널과 불일치합니다. 원본 영상의 변조(반전·마스킹) 흔적이 전처리 단계에서 탐지되었습니다.",
        "C5": "명확한 기획 의도와 고유한 편집 스타일이 확인되었습니다. 의미적 유사도·객체 일치도·싱크 점수 모두 정상 범위 내에 있어 우수 콘텐츠로 판정합니다.",
    }
    status_map = {"C1": "AUTO_REJECT", "C2": "AUTO_REJECT", "C3": "HUMAN_REVIEW", "C4": "AUTO_REJECT", "C5": "AUTO_APPROVE"}
    return {
        "video_id": video_id,
        "category": cat,
        "confidence_score": conf,
        "reasoning_log": reasoning_map[cat],
        "status": status_map[cat],
        "model_used": "gpt-4o-mini (서버 오프라인)",
        "processing_time": round(random.uniform(1.2, 4.8), 2),
        "context_score": ctx,
        "s_semantic": s,
        "o_existence": o,
        "a_sync": a,
        "layout_score": round(random.uniform(0.2, 0.9), 2),
        "raw_ocr_text": "[title_region] 🔥충격🔥 이것만 알면 100만원 [content_region] 돈버는법 클릭 지금바로 [ui_region] 좋아요 구독 알림",
        "channel_name": "Mock Channel",
        "view_count": random.randint(1000, 3000000),
        "duration": random.randint(15, 60),
        "created_at": datetime.now().isoformat(),
    }


def get_analysis_data(video_id: str) -> dict:
    """분석 데이터 가져오기: 세션 → API → Mock"""

    # 1. 세션 캐시
    if (st.session_state.get("latest_result") and
            st.session_state.latest_result.get("video_id") == video_id):
        r = st.session_state.latest_result
        td = r.get("technical_details", {})
        cs = r.get("context_score", {})   # API/직접호출 양쪽 경로 통합
        vi = r.get("video_info", {})
        return {
            "video_id": video_id,
            "category": r["analysis_result"]["category"],
            "confidence_score": r["analysis_result"]["confidence_score"],
            "reasoning_log": r["analysis_result"]["reasoning_log"],
            "status": r["analysis_result"]["status"],
            "model_used": r.get("model_used", "gpt-4o-mini"),
            "processing_time": r.get("processing_time", 1.5),
            "context_score":    cs.get("context_score", 0.75),
            "s_semantic":       cs.get("s_semantic", 0.8),
            "o_existence":      cs.get("o_existence", 0.7),
            "a_sync":           cs.get("a_sync", 0.8),
            # layout_score, keyframe_count: td(pipeline) → cs(직접호출) 순으로 fallback
            "layout_score":     td.get("layout_score") if td.get("layout_score") is not None else cs.get("layout_score", 0.0),
            "keyframe_count":   td.get("keyframe_count") if td.get("keyframe_count") is not None else cs.get("keyframe_count", 0),
            "raw_ocr_text":     td.get("ocr_text", "") or cs.get("ocr_text", ""),
            "ocr_text":         td.get("ocr_text", "") or cs.get("ocr_text", ""),
            "spam_detected":    td.get("spam_detected", False) or cs.get("spam_detected", False),
            "short_circuit_c4": td.get("short_circuit_c4", False) or cs.get("short_circuit_c4", False),
            # video_info 연동
            "channel_name": vi.get("channel", "—"),
            "view_count": vi.get("view_count", 0),
            "duration": vi.get("duration", 0),
            "created_at": datetime.now().isoformat(),
        }

    # 2. DB 연동 시도
    try:
        from database_manager import db_manager
        from database_models import AnalysisResults, Contents
        session = db_manager.get_session()
        try:
            analysis = session.query(AnalysisResults).join(Contents).filter(
                AnalysisResults.video_id == video_id
            ).order_by(AnalysisResults.created_at.desc()).first()
            if analysis and analysis.content:
                return {
                    "video_id": video_id,
                    "category": analysis.c_category,
                    "confidence_score": analysis.confidence_score,
                    "reasoning_log": analysis.reasoning_log,
                    "status": analysis.status.value,
                    "model_used": analysis.model_used,
                    "processing_time": analysis.processing_time,
                    "context_score": analysis.context_score or 0.75,
                    "s_semantic": analysis.s_semantic or 0.8,
                    "o_existence": analysis.o_existence or 0.7,
                    "a_sync": analysis.a_sync or 0.8,
                    "layout_score":     analysis.content.layout_score or 0.0,
                    "keyframe_count":   len(analysis.content.keyframes or []),
                    "raw_ocr_text":     analysis.content.raw_ocr_text or "",
                    "ocr_text":         analysis.content.raw_ocr_text or "",
                    "spam_detected":    False,
                    "short_circuit_c4": analysis.c_category in ["C1", "C2", "C3"],
                    "channel_name":     analysis.content.channel_name,
                    "view_count": analysis.content.view_count,
                    "duration": analysis.content.duration,
                    "created_at": analysis.created_at.isoformat(),
                }
        finally:
            session.close()
    except Exception:
        pass

    # 3. Mock fallback
    return make_mock_analysis(video_id)


def _save_analysis_to_db(result: dict, video_url: str) -> None:
    """분석 결과를 Contents + AnalysisResults 테이블에 저장 (직접 OpenAI 호출 경로용)"""
    try:
        import uuid as _uuid
        from database_manager import db_manager
        from database_models import Contents, AnalysisResults, ContentStatus, AnalysisResultStatus

        ar = result.get("analysis_result", {})
        cs = result.get("context_score", {})
        vi = result.get("video_info", {})
        video_id = result.get("video_id", "")

        session = db_manager.get_session()
        try:
            # Contents 없으면 생성
            content_row = session.query(Contents).filter(Contents.video_id == video_id).first()
            if not content_row:
                content_row = Contents(
                    video_id=video_id,
                    url=video_url,
                    title=vi.get("title", ""),
                    channel_name="",
                    duration=vi.get("duration") or vi.get("duration_sec"),
                    view_count=vi.get("view_count"),
                    layout_score=cs.get("layout_score", 0.0),
                    status=ContentStatus.COMPLETED,
                )
                session.add(content_row)
                session.flush()

            # AnalysisResults 저장
            status_map = {
                "AUTO_REJECT": AnalysisResultStatus.AUTO_REJECT,
                "HUMAN_REVIEW": AnalysisResultStatus.HUMAN_REVIEW,
                "AUTO_APPROVE": AnalysisResultStatus.AUTO_APPROVE,
            }
            analysis_row = AnalysisResults(
                result_id=f"result_{_uuid.uuid4().hex[:12]}",
                video_id=video_id,
                c_category=ar.get("category", "C5"),
                reasoning_log=ar.get("reasoning_log", ""),
                confidence_score=ar.get("confidence_score", 0.5),
                status=status_map.get(ar.get("status", "AUTO_APPROVE"), AnalysisResultStatus.AUTO_APPROVE),
                model_used=result.get("model_used", "gpt-4o-mini"),
                processing_time=result.get("processing_time", 0.0),
                context_score=cs.get("context_score"),
                s_semantic=cs.get("s_semantic"),
                o_existence=cs.get("o_existence"),
                a_sync=cs.get("a_sync"),
            )
            session.add(analysis_row)
            session.commit()
            logger.info(f"✅ 분석 결과 DB 저장 완료: {video_id}")
        finally:
            session.close()
    except Exception as e:
        logger.error(f"_save_analysis_to_db 실패: {e}")


def submit_feedback(video_id: str, action: str, text: str = "", source_url: str = "") -> bool:
    """피드백 저장: API 서버 우선 → 실패 시 DB 직접 저장.
    source_url: 실제 YouTube URL (관리자 테이블 링크용)
    """
    import uuid as _uuid

    # 1. API 서버 시도
    api_url = st.secrets.get("API_URL", None) or os.getenv("API_URL", "")
    if api_url:
        try:
            r = requests.post(
                f"{api_url.rstrip('/')}/feedback",
                json={"video_id": video_id, "action": action, "feedback_text": text},
                timeout=5
            )
            if r.status_code == 200:
                return True
        except Exception:
            pass

    # 2. DB 직접 저장 (API 없거나 실패 시)
    try:
        from database_manager import db_manager
        from database_models import UserFeedback
        session = db_manager.get_session()
        try:
            ctx = {}
            if source_url:
                ctx["source_url"] = source_url
            fb = UserFeedback(
                feedback_id=f"feedback_{_uuid.uuid4().hex[:8]}",
                video_id=video_id,
                user_action=action,
                feedback_type="user_action",
                feedback_text=text or None,
                feedback_context=ctx if ctx else None,
                is_processed=False,
            )
            session.add(fb)
            session.commit()
            return True
        finally:
            session.close()
    except Exception as e:
        logger.error(f"submit_feedback DB 저장 실패: {e}")
        return False



def submit_opinion_to_db(video_id: str, ai_category: str,
                          user_category: str, reason: str) -> bool:
    """의견을 API를 통해 DB에 저장"""
    feedback_text = f"[AI분류:{ai_category}→사용자제안:{user_category}] {reason}".strip()
    return submit_feedback(video_id, "opinion", feedback_text)

def check_api_health() -> bool:
    try:
        r = requests.get("http://localhost:8000/health", timeout=3)
        return r.status_code == 200
    except Exception:
        return False


# ─────────────────────────────────────────
# 차트 함수
# ─────────────────────────────────────────

def render_radar_chart(data: dict):
    fig = go.Figure()
    labels = ["의미적 유사도<br>(S_semantic)", "객체 존재<br>(O_existence)", "시공간 동기화<br>(A_sync)"]
    values = [
        data.get("s_semantic", 0) * 100,
        data.get("o_existence", 0) * 100,
        data.get("a_sync", 0) * 100,
    ]
    values_closed = values + [values[0]]
    labels_closed = labels + [labels[0]]

    # 배경 기준선
    for threshold, color in [(75, "rgba(16,185,129,0.08)"), (50, "rgba(245,158,11,0.06)")]:
        fig.add_trace(go.Scatterpolar(
            r=[threshold] * 4,
            theta=labels_closed,
            fill="toself",
            fillcolor=color,
            line=dict(color="rgba(255,255,255,0.04)", width=1),
            showlegend=False,
            hoverinfo="skip",
        ))

    fig.add_trace(go.Scatterpolar(
        r=values_closed,
        theta=labels_closed,
        fill="toself",
        fillcolor="rgba(124,131,253,0.18)",
        line=dict(color="#7c83fd", width=2.5),
        marker=dict(size=8, color="#7c83fd", symbol="circle"),
        showlegend=False,
        hovertemplate="%{theta}: %{r:.1f}%<extra></extra>",
    ))

    fig.update_layout(
        polar=dict(
            bgcolor="rgba(0,0,0,0)",
            radialaxis=dict(
                visible=True, range=[0, 100],
                ticksuffix="%", tickfont=dict(size=10, color="#999999"),
                gridcolor="rgba(0,0,0,0.10)",
                linecolor="rgba(0,0,0,0.12)",
                tickcolor="rgba(0,0,0,0.15)",
                showline=True,
                gridwidth=1,
            ),
            angularaxis=dict(
                tickfont=dict(size=11, color="#555555"),
                gridcolor="rgba(0,0,0,0.10)",
                linecolor="rgba(0,0,0,0.12)",
                gridwidth=1,
            ),
        ),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=0, r=0, t=20, b=20),
        height=440,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def render_confidence_gauge(confidence: float, category: str):
    verdict = get_verdict_class(category)
    color_map = {"danger": "#ef4444", "warning": "#f59e0b", "safe": "#10b981"}
    color = color_map.get(verdict, "#7c83fd")

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=confidence * 100,
        number={"suffix": "%", "font": {"size": 32, "color": color, "family": "Syne"}},
        title={"text": "AI 확신도", "font": {"size": 13, "color": "#aaaaaa", "family": "Syne"}},
        domain={"x": [0.05, 0.95], "y": [0.05, 0.95]},
        gauge={
            "axis": {"range": [0, 100], "tickwidth": 1, "tickcolor": "#999999",
                     "tickfont": {"size": 10, "color": "#999999"}},
            "bar": {"color": color, "thickness": 0.28},
            "bgcolor": "rgba(0,0,0,0)",
            "borderwidth": 0,
            "steps": [
                {"range": [0, 50],  "color": "rgba(239,68,68,0.08)"},
                {"range": [50, 80], "color": "rgba(245,158,11,0.08)"},
                {"range": [80, 100],"color": "rgba(16,185,129,0.08)"},
            ],
            "threshold": {"line": {"color": color, "width": 3}, "thickness": 0.8, "value": confidence * 100},
        },
    ))
    fig.update_layout(
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        margin=dict(l=20, r=20, t=80, b=40),
        height=440,
        font={"color": "#555555"},
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def render_category_bar():
    """카테고리 분포 가로 막대 차트"""
    cat_counts = safe_db_category_dist()

    colors = {"C1": "#ef4444", "C2": "#f97316", "C3": "#f59e0b", "C4": "#8b5cf6", "C5": "#10b981"}
    labels = [f"{k} {CATEGORY_META[k]['label']}" for k in cat_counts]
    values = list(cat_counts.values())
    bar_colors = [colors[k] for k in cat_counts]

    fig = go.Figure(go.Bar(
        x=values, y=labels, orientation="h",
        marker_color=bar_colors,
        marker=dict(cornerradius=6),
        hovertemplate="%{y}: %{x}건<extra></extra>",
        text=values, textposition="outside",
        textfont=dict(size=11, color="#555555"),
    ))
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(tickfont=dict(size=11, color="#555555"),
                   gridcolor="rgba(0,0,0,0.06)"),
        margin=dict(l=10, r=40, t=10, b=10),
        height=220,
        bargap=0.35,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


# ─────────────────────────────────────────
# 화면 렌더링: 분석 리포트
# ─────────────────────────────────────────

def show_report(video_id: str):
    # 리포트 화면에서는 메인 상단 여백 상쇄
    st.markdown("<div style='margin-top:-18vh'></div>", unsafe_allow_html=True)
    data = get_analysis_data(video_id)
    cat = data["category"]
    meta = CATEGORY_META.get(cat, CATEGORY_META["C5"])
    verdict = meta["verdict"]

    # 뒤로가기
    if st.button("← 메인으로 돌아가기", key="back_btn"):
        st.session_state.show_report = False
        st.session_state.video_id = None
        st.rerun()

    st.markdown("<div class='section-header'>분석 리포트</div>", unsafe_allow_html=True)

    # ── Verdict Card + 상세 정보 ──────────────────
    st.markdown("""
    <style>
      /* 행 전체 stretch */
      div[data-testid="stHorizontalBlock"]:has(.verdict-card) {
        align-items: stretch !important;
      }
      /* Streamlit 컬럼 내부 래퍼 전체 체인 height:100% */
      div[data-testid="stHorizontalBlock"]:has(.verdict-card)
        > div[data-testid="stVerticalBlockBorderWrapper"],
      div[data-testid="stHorizontalBlock"]:has(.verdict-card)
        > div[data-testid="stVerticalBlockBorderWrapper"] > div,
      div[data-testid="stHorizontalBlock"]:has(.verdict-card)
        > div[data-testid="stVerticalBlockBorderWrapper"] > div
        > div[data-testid="stVerticalBlock"],
      div[data-testid="stHorizontalBlock"]:has(.verdict-card)
        > div[data-testid="stVerticalBlockBorderWrapper"] > div
        > div[data-testid="stVerticalBlock"] > div:first-child {
        height: 100% !important;
        display: flex !important;
        flex-direction: column !important;
      }
      /* 카드 자체 */
      .verdict-card {
        flex: 1 !important;
        height: 100% !important;
        box-sizing: border-box !important;
      }
      /* 정보 패널 */
      .info-panel {
        flex: 1 !important;
        background: #f8f9fa;
        border: 1px solid #e9ecef;
        border-radius: 16px;
        padding: 16px 20px;
        height: 100%;
        box-sizing: border-box;
      }
    </style>
    """, unsafe_allow_html=True)
    col_card, col_info = st.columns([3, 2])

    with col_card:
        # ── 유튜브 영상 embed 준비 ──
        _url_for_embed = st.session_state.get("analyzed_url", "")
        _vid_id_embed  = _extract_video_id(_url_for_embed) if _url_for_embed and _url_for_embed != "—" else ""
        _has_video     = bool(_vid_id_embed and not _vid_id_embed.startswith("video_"))
        _thumb_url     = f"https://img.youtube.com/vi/{_vid_id_embed}/hqdefault.jpg" if _vid_id_embed else ""

        if _has_video:
            # 카테고리 정보(왼쪽) + 세로 영상(오른쪽) 나란히
            st.markdown(f"""
            <div class="verdict-card verdict-{verdict}"
                 style="display:flex;align-items:stretch;gap:20px;padding:20px;">

              <!-- 왼쪽: 카테고리 정보 -->
              <div style="flex:1;display:flex;flex-direction:column;justify-content:center;gap:12px;">
                <div style="font-size:3rem;line-height:1">{meta['icon']}</div>
                <div>
                  <p class="verdict-category"
                     style="color:{'#fca5a5' if verdict=='danger' else '#fcd34d' if verdict=='warning' else '#6ee7b7'};
                            margin:0 0 4px 0">{cat}</p>
                  <p class="verdict-label" style="margin:0 0 8px 0">{meta['label']}</p>
                  <span class="verdict-badge badge-{verdict}">{meta['badge']}</span>
                </div>
              </div>

              <!-- 오른쪽: 세로형 숏츠 (9:16) -->
              <div style="width:160px;flex-shrink:0;">
                <div style="position:relative;width:160px;height:284px;
                            border-radius:12px;overflow:hidden;background:#000;">
                  <iframe
                    src="https://www.youtube.com/embed/{_vid_id_embed}?rel=0&modestbranding=1"
                    style="position:absolute;top:0;left:0;width:100%;height:100%;border:none;"
                    allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture"
                    allowfullscreen>
                  </iframe>
                </div>
              </div>

            </div>
            """, unsafe_allow_html=True)
        else:
            # 영상 없을 때 기존 레이아웃 + 썸네일
            st.markdown(f"""
            <div class="verdict-card verdict-{verdict}">
              <div style="display:flex;align-items:flex-start;gap:18px;">
                <div style="font-size:3.5rem;line-height:1">{meta['icon']}</div>
                <div>
                  <p class="verdict-category" style="color:{'#fca5a5' if verdict=='danger' else '#fcd34d' if verdict=='warning' else '#6ee7b7'}">{cat}</p>
                  <p class="verdict-label">{meta['label']}</p>
                  <span class="verdict-badge badge-{verdict}">{meta['badge']}</span>
                </div>
              </div>
              {f'<img src="{_thumb_url}" style="width:100%;border-radius:10px;margin-top:12px;" onerror="this.style.display=\'none\'">' if _thumb_url else ''}
            </div>
            """, unsafe_allow_html=True)

    with col_info:
        _url_val   = st.session_state.get("analyzed_url", "—")
        _title_val = st.session_state.get("analyzed_title", "—")

        # 조회수: 세션 우선, 없으면 data 폴백
        _views_raw = st.session_state.get("analyzed_views")
        if _views_raw is not None:
            _views_str = f"{_views_raw:,}회"
        else:
            _v = data.get('view_count', 0)
            _views_str = f"{_v:,}회" if _v else "—"

        # 길이: 세션 우선, 없으면 data 폴백 → mm:ss 변환
        _dur_raw = st.session_state.get("analyzed_duration")
        if _dur_raw is not None:
            _dur_str = f"{_dur_raw // 60}:{_dur_raw % 60:02d}" if _dur_raw >= 60 else f"{_dur_raw}초"
        else:
            _d = data.get('duration', 0)
            _dur_str = f"{_d}초" if _d else "—"

        _info_items = [
            ("URL",    _url_val),
            ("제목",   _title_val),
            ("조회수", _views_str),
            ("길이",   _dur_str),
            ("모델",   data.get('model_used', '—')),
            ("처리 시간", f"{data.get('processing_time', 0):.2f}s"),
        ]
        info_html = "<div class='info-panel'>"
        for _label, _value in _info_items:
            _font = "color:#888;font-size:0.75rem" if _label in ("URL", "제목") else "color:#444;font-size:0.88rem;font-weight:500"
            info_html += (
                f"<div style='display:flex;align-items:baseline;gap:10px;padding:7px 0;"
                f"border-bottom:1px solid #f4f4f4;'>"
                f"<span style='color:#3d4263;font-size:0.72rem;text-transform:uppercase;"
                f"letter-spacing:0.08em;min-width:68px;flex-shrink:0'>{_label}</span>"
                f"<span style='{_font};word-break:break-all'>{_value}</span></div>"
            )
        info_html += "</div>"
        st.markdown(info_html, unsafe_allow_html=True)

    # ── AI 판단 근거 ──────────────────────────
    st.markdown("<div class='section-header'>AI 판단 근거 (Reasoning View)</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='reasoning-box'>💬 {data.get('reasoning_log','—')}</div>", unsafe_allow_html=True)

    # ── CIS + 분류별 위험도 ──────────────────────────
    st.markdown("<div class='section-header'>CIS 분석 및 분류별 위험도</div>", unsafe_allow_html=True)

    # C1/C2/C3 스코어 계산
    confidence = data.get("confidence_score", 0.5)
    context = data.get("context_score", 0.75)

    def _cat_score(target):
        if cat == target:
            return round(confidence * 100, 1)
        elif cat == "C5":
            return round((1 - confidence) * 30, 1)
        else:
            return round((1 - confidence) * 50, 1)

    score_c1 = data.get("score_c1", _cat_score("C1"))
    score_c2 = data.get("score_c2", _cat_score("C2"))
    score_c3 = data.get("score_c3", _cat_score("C3"))

    # CIS 계산
    alpha, beta = 0.3, 0.3
    cis = max(0.0, min(1.0, context - (alpha * score_c1/100 + beta * score_c2/100)))

    # 카테고리별 색상
    color_map = {"C1":"#ef4444","C2":"#f97316","C3":"#f59e0b","C4":"#8b5cf6","C5":"#10b981"}
    fill_map  = {"C1":"rgba(239,68,68,0.18)","C2":"rgba(249,115,22,0.18)",
                 "C3":"rgba(245,158,11,0.18)","C4":"rgba(139,92,246,0.18)","C5":"rgba(16,185,129,0.18)"}
    verdict   = get_verdict_class(cat)
    v_color   = {"danger":"#ef4444","warning":"#f59e0b","safe":"#10b981"}.get(verdict,"#7c83fd")
    line_color = color_map.get(cat, "#7c83fd")
    fill_color = fill_map.get(cat, "rgba(124,131,253,0.18)")

    # ── 한 줄 4컬럼: C1레이더 | C2레이더 | C3레이더 | CIS 게이지 ──
    col_c1, col_c2, col_c3, col_cis = st.columns([1, 1, 1, 1])

    def _make_radar(title, score, sub_labels, sub_values, line_col, fill_col):
        """개별 레이더 차트 생성"""
        vals_c = sub_values + [sub_values[0]]
        lbls_c = sub_labels + [sub_labels[0]]
        fig = go.Figure()
        for thr, clr in [(80,"rgba(239,68,68,0.08)"),(50,"rgba(245,158,11,0.06)")]:
            fig.add_trace(go.Scatterpolar(
                r=[thr]*( len(sub_labels)+1), theta=lbls_c,
                fill="toself", fillcolor=clr,
                line=dict(color="rgba(0,0,0,0.04)", width=1),
                showlegend=False, hoverinfo="skip"))
        fig.add_trace(go.Scatterpolar(
            r=vals_c, theta=lbls_c, fill="toself",
            fillcolor=fill_col, line=dict(color=line_col, width=2.5),
            marker=dict(size=7, color=line_col),
            showlegend=False, hovertemplate="%{theta}: %{r:.1f}<extra></extra>"))
        fig.update_layout(
            polar=dict(
                bgcolor="rgba(0,0,0,0)",
                radialaxis=dict(visible=True, range=[0,100],
                    tickfont=dict(size=8, color="#aaa"),
                    gridcolor="rgba(0,0,0,0.08)", linecolor="rgba(0,0,0,0.10)"),
                angularaxis=dict(tickfont=dict(size=9, color="#666"),
                    gridcolor="rgba(0,0,0,0.08)")),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=10, r=10, t=10, b=10), height=220)
        return fig

    # C1 레이더 — 어그로/스팸
    with col_c1:
        st.markdown("""
        <div style='text-align:center;font-size:0.72rem;font-weight:700;
                    color:#ef4444;letter-spacing:0.06em;margin-bottom:4px'>
            C1 · 어그로/스팸
        </div>""", unsafe_allow_html=True)
        c1_labels = ["키워드<br>위험도", "의미론적<br>유사도", "클릭베이트<br>패턴"]
        c1_vals   = [
            round(score_c1 * 1.0, 1),
            round(score_c1 * 0.85, 1),
            round(score_c1 * 0.9, 1),
        ]
        st.plotly_chart(
            _make_radar("C1", score_c1, c1_labels, c1_vals, "#ef4444", "rgba(239,68,68,0.18)"),
            use_container_width=True, config={"displayModeBar": False}
        )
        st.markdown(f"<div style='text-align:center;font-size:0.8rem;font-weight:700;color:#ef4444'>{score_c1:.1f} / 100</div>", unsafe_allow_html=True)

    # C2 레이더 — 공장형 패턴 (의미적 유사도 기반)
    with col_c2:
        st.markdown("""
        <div style='text-align:center;font-size:0.72rem;font-weight:700;
                    color:#f97316;letter-spacing:0.06em;margin-bottom:4px'>
            C2 · 공장형 패턴
        </div>""", unsafe_allow_html=True)
        s_val     = data.get("s_semantic", 0)
        c2_labels = ["의미적<br>유사도", "레이아웃<br>반복성", "TTS<br>패턴"]
        c2_vals   = [
            round(s_val * 100, 1),
            round(score_c2 * 0.9, 1),
            round(score_c2 * 0.8, 1),
        ]
        st.plotly_chart(
            _make_radar("C2", score_c2, c2_labels, c2_vals, "#f97316", "rgba(249,115,22,0.18)"),
            use_container_width=True, config={"displayModeBar": False}
        )
        st.markdown(f"<div style='text-align:center;font-size:0.8rem;font-weight:700;color:#f97316'>{score_c2:.1f} / 100</div>", unsafe_allow_html=True)

    # C3 레이더 — 품질 불량 (시공간 동기화 기반)
    with col_c3:
        st.markdown("""
        <div style='text-align:center;font-size:0.72rem;font-weight:700;
                    color:#f59e0b;letter-spacing:0.06em;margin-bottom:4px'>
            C3 · 품질 불량
        </div>""", unsafe_allow_html=True)
        a_val     = data.get("a_sync", 0)
        c3_labels = ["시공간<br>동기화", "자막-화면<br>일치도", "품질<br>점수"]
        c3_vals   = [
            round(a_val * 100, 1),
            round(score_c3 * 0.9, 1),
            round(score_c3 * 0.85, 1),
        ]
        st.plotly_chart(
            _make_radar("C3", score_c3, c3_labels, c3_vals, "#f59e0b", "rgba(245,158,11,0.18)"),
            use_container_width=True, config={"displayModeBar": False}
        )
        st.markdown(f"<div style='text-align:center;font-size:0.8rem;font-weight:700;color:#f59e0b'>{score_c3:.1f} / 100</div>", unsafe_allow_html=True)

    # CIS 게이지
    with col_cis:
        st.markdown("<div style='text-align:center;font-size:0.72rem;color:#aaa;margin-bottom:4px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em'>CIS Score</div>", unsafe_allow_html=True)
        fig_cis = go.Figure(go.Indicator(
            mode="gauge+number",
            value=round(cis * 100, 1),
            number={"font":{"size":28,"color":v_color,"family":"Syne"},"valueformat":".1f"},
            title={"text":"CIS_Final","font":{"size":11,"color":"#aaaaaa","family":"Syne"}},
            domain={"x":[0.05,0.95],"y":[0.05,0.95]},
            gauge={
                "axis":{"range":[0,100],"tickfont":{"size":9,"color":"#bbb"}},
                "bar":{"color":v_color,"thickness":0.28},
                "bgcolor":"rgba(0,0,0,0)", "borderwidth":0,
                "steps":[
                    {"range":[0,40],  "color":"rgba(239,68,68,0.08)"},
                    {"range":[40,70], "color":"rgba(245,158,11,0.08)"},
                    {"range":[70,100],"color":"rgba(16,185,129,0.08)"},
                ],
                "threshold":{"line":{"color":v_color,"width":3},"thickness":0.8,"value":cis*100},
            }))
        fig_cis.update_layout(
            paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
            margin=dict(l=10,r=10,t=50,b=10), height=200)
        st.plotly_chart(fig_cis, use_container_width=True, config={"displayModeBar":False})
        st.markdown(f"""
        <div style="text-align:center;font-size:0.72rem;color:#aaa;margin-top:-8px">
            Context − (α·C1 + β·C2)<br>
            <span style="color:{v_color};font-weight:700">
            {"정상 ✅" if cis>=0.7 else "검토 필요 ⚠️" if cis>=0.4 else "위험 🚨"}
            </span>
        </div>
        """, unsafe_allow_html=True)

    # ── 사용자 액션 ───────────────────────────
    st.markdown("""
    <div style="
        background:#f8f9fa;
        border:1px solid #e9ecef;
        border-radius:14px;
        padding:18px 24px;
        text-align:center;
        margin:28px 0 14px 0;
    ">
      <span style="
        font-family:'DM Sans',sans-serif;
        font-size:0.95rem;
        font-weight:500;
        color:#555555;
        letter-spacing:0.01em;
      ">이 영상에 대해 어떻게 생각하시나요?</span>
    </div>
    """, unsafe_allow_html=True)

    # 세션 초기화
    if "show_opinion_form" not in st.session_state:
        st.session_state.show_opinion_form = False
    if "show_report_form" not in st.session_state:
        st.session_state.show_report_form = False
    if "action_result" not in st.session_state:
        st.session_state.action_result = None

    # 유튜브 링크 생성 헬퍼
    yt_url       = f"https://www.youtube.com/watch?v={video_id}"
    yt_report_url = f"https://www.youtube.com/watch?v={video_id}&action=flag"
    channel_name  = data.get("channel_name", "")

    # 채널 페이지: channel_name이 있으면 검색, 없으면 영상 페이지
    yt_channel_url = (
        f"https://www.youtube.com/results?search_query={requests.utils.quote(channel_name)}"
        if channel_name and channel_name not in ("—", "Mock Channel")
        else yt_url
    )

    # 버튼 5개 (좋아요, 싫어요, 채널추천안함, 신고하기, 의견보내기)
    btn_cols = st.columns(5)
    with btn_cols[0]:
        if st.button("👍 좋아요", use_container_width=True, key="act_like"):
            submit_feedback(video_id, "like", source_url=st.session_state.get("analyzed_url",""))
            st.session_state.action_result = (
                "success",
                f"좋아요가 기록되었습니다! 유튜브에서도 좋아요를 누르려면 → [영상 바로가기]({yt_url})"
            )
            st.session_state.show_opinion_form = False
    with btn_cols[1]:
        if st.button("👎 싫어요", use_container_width=True, key="act_dislike"):
            submit_feedback(video_id, "dislike", source_url=st.session_state.get("analyzed_url",""))
            st.session_state.action_result = (
                "success",
                f"싫어요가 기록되었습니다! 유튜브에서도 싫어요를 누르려면 → [영상 바로가기]({yt_url})"
            )
            st.session_state.show_opinion_form = False
    with btn_cols[2]:
        if st.button("🚫 채널 추천 안 함", use_container_width=True, key="act_block"):
            submit_feedback(video_id, "block_channel", source_url=st.session_state.get("analyzed_url",""))
            st.session_state.action_result = (
                "success",
                f"채널 추천 안 함이 기록되었습니다! 유튜브에서 채널을 차단하려면 → [채널 바로가기]({yt_channel_url})"
            )
            st.session_state.show_opinion_form = False
    with btn_cols[3]:
        report_label = "📢 신고하기 ▲" if st.session_state.show_report_form else "📢 신고하기 ▼"
        if st.button(report_label, use_container_width=True, key="act_report"):
            st.session_state.show_report_form = not st.session_state.show_report_form
            st.session_state.show_opinion_form = False
            st.session_state.action_result = None
    with btn_cols[4]:
        btn_label = "✏️ 의견 보내기 ▲" if st.session_state.show_opinion_form else "✏️ 의견 보내기 ▼"
        if st.button(btn_label, use_container_width=True, key="act_opinion"):
            st.session_state.show_opinion_form = not st.session_state.show_opinion_form
            st.session_state.show_report_form = False
            st.session_state.action_result = None

    # 액션 결과 표시
    if st.session_state.action_result:
        msg_type, msg_text = st.session_state.action_result
        if msg_type == "success":
            st.success(msg_text)
        elif msg_type == "error":
            st.error(msg_text)


    # ── 신고하기 확장 폼 ─────────────────────────
    if st.session_state.show_report_form:
        REPORT_CATEGORIES = [
            ("🔞", "성적인 콘텐츠"),
            ("💀", "폭력적 또는 혐오스러운 콘텐츠"),
            ("😡", "증오 또는 악의적인 콘텐츠"),
            ("😰", "괴롭힘 또는 폭력"),
            ("⚠️", "유해하거나 위험한 행위"),
            ("🆘", "자살, 자해 또는 섭식 장애"),
            ("❌", "잘못된 정보"),
            ("👶", "아동 학대"),
            ("💣", "테러 조장"),
            ("🚫", "스팸 또는 혼동을 야기하는 콘텐츠"),
            ("⚖️", "법적 문제"),
        ]

        if "selected_report" not in st.session_state:
            st.session_state.selected_report = None

        # 선택 상태에 따라 pill 버튼 CSS를 동적으로 생성
        pill_css = ""
        for _, lbl in REPORT_CATEGORIES:
            key = f"rpt_{lbl}"
            if st.session_state.selected_report == lbl:
                pill_css += f"""
                div[data-testid="stButton"]:has(button[data-testid="{key}"] ),
                div[data-testid="stButton"] button[key="{key}"] {{
                    background: #cc0000 !important;
                    color: #ffffff !important;
                    border-color: #cc0000 !important;
                }}"""

        st.markdown(f"""
        <style>
          /* 신고 pill 버튼 공통: 둥글게 */
          .report-pill-area div[data-testid="stButton"] > button {{
            border-radius: 50px !important;
            font-size: 0.82rem !important;
            font-weight: 600 !important;
            padding: 8px 10px !important;
            transition: all 0.15s ease !important;
            border: 1.5px solid #dddddd !important;
            background: #ffffff !important;
            color: #333333 !important;
            min-height: 44px !important;
          }}
          .report-pill-area div[data-testid="stButton"] > button:hover {{
            border-color: #cc0000 !important;
            color: #cc0000 !important;
          }}
        </style>
        <div style="background:#fff8f8;border:1.5px solid rgba(204,0,0,0.18);
                    border-radius:14px;padding:20px 24px 4px 24px;margin-top:12px;">
          <p style="font-family:'Syne',sans-serif;font-size:1rem;font-weight:700;
                    color:#cc0000;margin:0 0 2px 0">📢 신고하기</p>
          <p style="font-size:0.85rem;color:#666;margin:0 0 14px 0">어떤 문제인가요?</p>
        </div>
        """, unsafe_allow_html=True)

        # pill 버튼 영역 (클래스로 감싸서 CSS 타겟팅)
        st.markdown('<div class="report-pill-area">', unsafe_allow_html=True)

        items_per_row = 3
        rows = [REPORT_CATEGORIES[i:i+items_per_row]
                for i in range(0, len(REPORT_CATEGORIES), items_per_row)]

        for row in rows:
            r_cols = st.columns(items_per_row)
            for idx, (icon, lbl) in enumerate(row):
                with r_cols[idx]:
                    is_sel = st.session_state.selected_report == lbl
                    # 선택됐으면 버튼 스타일 인라인으로 주입
                    if is_sel:
                        st.markdown(f"""
                        <style>
                          div[data-testid="stButton"]:has(> button p)
                          + div {{ display:none }}
                        </style>
                        """, unsafe_allow_html=True)
                    btn_text = f"{'✓ ' if is_sel else ''}{icon} {lbl}"
                    if st.button(btn_text, key=f"rpt_{lbl}", use_container_width=True):
                        st.session_state.selected_report = None if is_sel else lbl
                        st.rerun()
                    # 선택 항목은 CSS로 빨간색 덮어쓰기
                    if is_sel:
                        st.markdown(f"""
                        <style>
                          div[data-testid="stButton"]:has(button:contains("✓")) > button {{
                            background: #cc0000 !important;
                            color: #ffffff !important;
                            border-color: #cc0000 !important;
                            box-shadow: 0 2px 8px rgba(204,0,0,0.3) !important;
                          }}
                        </style>
                        """, unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

        sub_col, can_col = st.columns(2)
        with sub_col:
            can_submit = st.session_state.selected_report is not None
            if st.button("📤 신고 제출하기", key="report_submit",
                         use_container_width=True, disabled=not can_submit):
                selected = st.session_state.selected_report
                submit_feedback(video_id, "report", f"[신고 유형: {selected}]", source_url=st.session_state.get("analyzed_url",""))
                st.session_state.action_result = (
                    "success",
                    f"✅ '{selected}' 유형으로 신고가 접수되었습니다! "
                    f"유튜브에도 직접 신고하려면 → [유튜브 신고 페이지]({yt_report_url})"
                )
                st.session_state.show_report_form = False
                st.session_state.selected_report = None
                st.rerun()
        with can_col:
            if st.button("✕ 취소", key="report_cancel", use_container_width=True):
                st.session_state.show_report_form = False
                st.session_state.selected_report = None
                st.rerun()

        if not st.session_state.selected_report:
            st.caption("※ 신고 유형을 하나 선택해주세요.")


    # ── 의견 보내기 확장 폼 ──────────────────────
    if st.session_state.show_opinion_form:
        st.markdown("""
        <div style="background:#fff8f8;border:1.5px solid rgba(204,0,0,0.15);
                    border-radius:14px;padding:24px 28px;margin-top:12px">
          <div style="font-family:'Syne',sans-serif;font-size:0.8rem;font-weight:700;
                      letter-spacing:0.1em;text-transform:uppercase;color:#cc0000;
                      margin-bottom:16px">✏️ 분류 의견 제출</div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("opinion_form", clear_on_submit=True):
            st.markdown(
                "<p style='font-size:0.85rem;color:#666;margin-bottom:4px'>"
                f"AI 분류 결과: <b style='color:#cc0000'>{cat} — {CATEGORY_META[cat]['label']}</b></p>",
                unsafe_allow_html=True
            )

            # 카테고리 직접 선택
            cat_options = {
                "C1 — 어그로 / 스팸": "C1",
                "C2 — 공장형 패턴":   "C2",
                "C3 — 품질 불량":     "C3",
                "C4 — 무단 도용":     "C4",
                "C5 — 정상 영상":     "C5",
            }
            selected_label = st.selectbox(
                "올바른 분류라고 생각하시는 카테고리를 선택해주세요",
                options=list(cat_options.keys()),
                index=list(cat_options.values()).index(cat),
                key="opinion_category"
            )
            user_category = cat_options[selected_label]

            # 이유 작성
            reason = st.text_area(
                "이유를 작성해주세요 (선택)",
                placeholder="예: 영상 내용을 직접 봤는데 실제로는 정상적인 교육 콘텐츠였습니다.",
                height=100,
                key="opinion_reason"
            )

            col_submit, col_cancel = st.columns([1, 1])
            with col_submit:
                submitted_opinion = st.form_submit_button("📤 제출하기", use_container_width=True)
            with col_cancel:
                cancel = st.form_submit_button("✕ 취소", use_container_width=True)

            if submitted_opinion:
                if not reason.strip() and user_category == cat:
                    st.warning("AI 분류 결과와 동일하거나 이유를 입력해주세요.")
                else:
                    # DB에 저장 (action=opinion, feedback_text에 AI분류→사용자제안+이유 포함)
                    saved = submit_opinion_to_db(
                        video_id=video_id,
                        ai_category=f"{cat}({CATEGORY_META[cat]['label']})",
                        user_category=f"{user_category}({CATEGORY_META[user_category]['label']})",
                        reason=reason or "(이유 미작성)"
                    )
                    if saved:
                        st.session_state.action_result = ("success", "의견이 DB에 저장되었습니다! 관리자가 검토할 예정입니다.")
                    else:
                        st.session_state.action_result = ("error", "저장 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.")

                    st.session_state.show_opinion_form = False
                    st.rerun()

            if cancel:
                st.session_state.show_opinion_form = False
                st.rerun()

    # ── 기술 세부사항 ──────────────────────────
    with st.expander("🔬 기술적 세부사항", expanded=False):

        # ── 1. CIS 계산 상세 내역 ──
        st.markdown("<div style='font-size:0.78rem;font-weight:700;color:#3d4263;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px'>① CIS 계산 상세</div>", unsafe_allow_html=True)

        confidence = data.get("confidence_score", 0.5)
        context    = data.get("context_score", 0.75)
        cat_now    = data.get("category", "C5")

        def _cat_score_detail(target):
            if cat_now == target:
                return round(confidence * 100, 1)
            elif cat_now == "C5":
                return round((1 - confidence) * 30, 1)
            else:
                return round((1 - confidence) * 50, 1)

        sc1 = data.get("score_c1", _cat_score_detail("C1"))
        sc2 = data.get("score_c2", _cat_score_detail("C2"))
        alpha, beta = 0.3, 0.3
        penalty = round(alpha * sc1/100 + beta * sc2/100, 4)
        cis_val = round(max(0.0, min(1.0, context - penalty)), 4)

        cis_color = "#10b981" if cis_val >= 0.7 else "#f59e0b" if cis_val >= 0.4 else "#ef4444"
        st.markdown(f"""
        <div style="background:#f8f9fa;border-radius:10px;padding:16px 20px;
                    font-family:monospace;font-size:0.85rem;line-height:2">
            <div>Context_Score (C3)  =  <b>{context:.4f}</b></div>
            <div>α · Score_C1        =  {alpha} × {sc1:.1f}/100  =  <b>{alpha * sc1/100:.4f}</b></div>
            <div>β · Score_C2        =  {beta} × {sc2:.1f}/100  =  <b>{beta * sc2/100:.4f}</b></div>
            <div style="border-top:1px solid #dee2e6;margin-top:6px;padding-top:6px">
                CIS_Final = {context:.4f} − {penalty:.4f}
                = <span style="color:{cis_color};font-weight:700;font-size:1rem">{cis_val:.4f}</span>
                &nbsp;{"✅ 정상" if cis_val >= 0.7 else "⚠️ 검토 필요" if cis_val >= 0.4 else "🚨 위험"}
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # ── 2. 분류 판단 단계별 로그 ──
        st.markdown("<div style='font-size:0.78rem;font-weight:700;color:#3d4263;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px'>② 분류 판단 단계별 로그</div>", unsafe_allow_html=True)

        spam_detected   = data.get("spam_detected", False)
        short_circuit   = data.get("short_circuit_c4", cat_now in ["C1","C2","C3"])
        reasoning_full  = data.get("reasoning_log", "—")

        step1_color  = "#10b981"
        step1_result = "피사체 움직임 감지 → C2 제외 조건 확인"

        step2_color  = "#10b981" if not spam_detected else "#ef4444"
        step2_result = "스팸 패턴 감지 → C1 즉시 분류 (API 호출 생략)" if spam_detected else "스팸 패턴 없음 → GPT-4o 분석 진행"

        step3_color  = "#10b981" if not short_circuit else "#f59e0b"
        step3_result = f"C4 검사 건너뜀 (Short-circuit) — {cat_now} 확정" if short_circuit else "C4 무단 도용 검사 수행"

        step4_color  = {"C1":"#ef4444","C2":"#f97316","C3":"#f59e0b","C4":"#8b5cf6","C5":"#10b981"}.get(cat_now,"#888")
        step4_result = f"최종 판정: {cat_now} — {CATEGORY_META.get(cat_now,{}).get('label','—')}"

        for step_num, step_name, step_res, s_color in [
            ("STEP 1", "피사체 움직임 확인",     step1_result, step1_color),
            ("STEP 2", "스팸 패턴 사전 체크",    step2_result, step2_color),
            ("STEP 3", "C4 Short-circuit 판단", step3_result, step3_color),
            ("STEP 4", "최종 분류 결정",         step4_result, step4_color),
        ]:
            st.markdown(f"""
            <div style="display:flex;align-items:flex-start;gap:12px;
                        padding:10px 0;border-bottom:1px solid #f0f0f0">
                <span style="background:{s_color};color:white;font-size:0.68rem;
                             font-weight:700;padding:3px 8px;border-radius:20px;
                             white-space:nowrap;margin-top:2px">{step_num}</span>
                <div>
                    <div style="font-size:0.8rem;font-weight:600;color:#333">{step_name}</div>
                    <div style="font-size:0.78rem;color:#666;margin-top:2px">{step_res}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # ── 3. 키프레임 정보 ──
        st.markdown("<div style='font-size:0.78rem;font-weight:700;color:#3d4263;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px'>③ 키프레임 정보</div>", unsafe_allow_html=True)

        keyframe_count = data.get("keyframe_count", data.get("processing_time", 0))
        layout_score   = data.get("layout_score", 0)
        analysis_time  = data.get("processing_time", 0)

        # 직접 OpenAI 호출 경로는 프레임 추출이 없으므로 0이면 N/A 표시
        kf_display     = f"{int(keyframe_count)}장" if keyframe_count else "N/A (메타데이터 분석)"
        layout_display = f"{layout_score:.3f}"       if layout_score  else "N/A (메타데이터 분석)"

        kf_cols = st.columns(4)
        for col, label, val in [
            (kf_cols[0], "추출 프레임 수",  kf_display),
            (kf_cols[1], "레이아웃 점수",   layout_display),
            (kf_cols[2], "분석 소요 시간",  f"{analysis_time:.2f}s"),
            (kf_cols[3], "사용 모델",       data.get("model_used", "gpt-4o-mini")),
        ]:
            with col:
                st.markdown(f"""
                <div style="background:#f8f9fa;border-radius:8px;padding:12px;text-align:center">
                    <div style="font-size:1rem;font-weight:700;color:#333">{val}</div>
                    <div style="font-size:0.72rem;color:#888;margin-top:4px">{label}</div>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        # ── 4. 스팸 패턴 매칭 결과 ──
        st.markdown("<div style='font-size:0.78rem;font-weight:700;color:#3d4263;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px'>④ 스팸 패턴 매칭</div>", unsafe_allow_html=True)
        reasoning_text = data.get("reasoning_log", "")
        if spam_detected and "감지" in reasoning_text:
            lines = [l.strip() for l in reasoning_text.split("\n") if l.strip()]
            for line in lines:
                color = "#ef4444" if "감지" in line else "#666"
                st.markdown(f"<div style='font-size:0.82rem;color:{color};padding:3px 0'>{line}</div>", unsafe_allow_html=True)
        elif cat_now == "C1":
            st.markdown("<div style='font-size:0.82rem;color:#ef4444'>⚠️ C1 관련 패턴 감지 (GPT 판단)</div>", unsafe_allow_html=True)
        else:
            st.markdown("<div style='font-size:0.82rem;color:#10b981'>✅ 스팸 패턴 없음</div>", unsafe_allow_html=True)


# ─────────────────────────────────────────
# 화면 렌더링: 메인 대시보드
# ─────────────────────────────────────────

def show_main():
    # ── 검색바 pill CSS (전역 stForm > div에 직접 타겟) ──
    st.markdown("""
    <style>
      /* form 자체 초기화 */
      div[data-testid="stForm"] {
        border: none !important;
        padding: 0 !important;
        background: transparent !important;
        box-shadow: none !important;
      }
      /* stForm 바로 아래 첫번째 div = 실질적인 컨테이너 */
      div[data-testid="stForm"] > div:first-child {
        display: flex !important;
        flex-direction: row !important;
        align-items: center !important;
        border: 1.8px solid #bbbbbb !important;
        border-radius: 999px !important;
        background: #ffffff !important;
        box-shadow: 0 2px 12px rgba(0,0,0,0.09) !important;
        overflow: hidden !important;
        padding: 0 !important;
        gap: 0 !important;
        transition: border-color 0.2s, box-shadow 0.2s, transform 0.2s !important;
      }
      div[data-testid="stForm"] > div:first-child:hover {
        border-color: #888 !important;
        box-shadow: 0 6px 24px rgba(0,0,0,0.14) !important;
        transform: translateY(-2px) !important;
      }
      div[data-testid="stForm"] > div:first-child:focus-within {
        border-color: #cc0000 !important;
        box-shadow: 0 6px 28px rgba(204,0,0,0.16) !important;
        transform: translateY(-2px) !important;
      }
      /* 입력창 */
      div[data-testid="stForm"] .stTextInput > div > div > input {
        border: none !important;
        box-shadow: none !important;
        background: transparent !important;
        height: 52px !important;
        padding-left: 24px !important;
        font-size: 0.96rem !important;
        color: #333 !important;
        border-radius: 0 !important;
      }
      div[data-testid="stForm"] .stTextInput > div > div,
      div[data-testid="stForm"] .stTextInput > div {
        border: none !important;
        background: transparent !important;
        box-shadow: none !important;
      }
      /* CHECK 버튼 */
      div[data-testid="stForm"] [data-testid="stFormSubmitButton"] > button {
        height: 52px !important;
        border-radius: 0 999px 999px 0 !important;
        border: none !important;
        border-left: 1.5px solid #e0e0e0 !important;
        padding: 0 28px !important;
        font-size: 0.88rem !important;
        font-weight: 700 !important;
        letter-spacing: 0.04em !important;
        background: #ffffff !important;
        color: #cc0000 !important;
        white-space: nowrap !important;
        transition: background 0.18s, color 0.18s !important;
        box-shadow: none !important;
      }
      div[data-testid="stForm"] [data-testid="stFormSubmitButton"] > button:hover {
        background: #cc0000 !important;
        color: #ffffff !important;
        border-left-color: #cc0000 !important;
        transform: none !important;
      }
    </style>
    """, unsafe_allow_html=True)

    # ── 검색바 중앙 정렬 ──────────────────────────
    _, center_col, _ = st.columns([1, 3, 1])
    with center_col:
        with st.form("analyze_form"):
            url_col, btn_col = st.columns([5, 1])
            with url_col:
                video_url = st.text_input(
                    "url", label_visibility="collapsed",
                    placeholder="YouTube Shorts URL을 붙여넣으세요  (예: https://youtube.com/shorts/...)"
                )
            with btn_col:
                submitted = st.form_submit_button("CHECK", use_container_width=True)

    if submitted and video_url.strip():
        raw_url = video_url.strip()

        # ── URL 형식 사전 검증 ─────────────────────────
        if not _is_valid_youtube_url(raw_url):
            _show_url_error("invalid", raw_url)
        else:
            # ── 영상 실제 존재 여부 확인 (oEmbed 404 체크) ──
            with st.spinner("🔍 영상 확인 중..."):
                video_exists = _check_video_exists(raw_url)
            if not video_exists:
                _show_url_error("not_found", raw_url)
            else:
                import time as _time
                _analysis_start = _time.time()
                TIMEOUT_SEC = 65   # 65초

                with st.spinner("🤖 AI 분석 중... (최대 65초)"):
                    try:
                        vid = _extract_video_id(raw_url)

                        # 1. API 서버 먼저 시도
                        api_url = (
                            st.secrets.get("API_URL", None)
                            or os.getenv("API_URL", "")
                        )
                        result = None

                        if api_url:
                            try:
                                r = requests.post(
                                    f"{api_url.rstrip('/')}/analyze",
                                    json={"video_url": raw_url, "request_source": "streamlit"},
                                    timeout=65
                                )
                                if r.status_code == 200:
                                    result = r.json()
                                elif r.status_code == 404:
                                    _show_url_error("not_found", raw_url)
                                    result = None
                            except requests.exceptions.Timeout:
                                pass  # 아래 5분 체크에서 처리
                            except Exception:
                                pass

                        # 2. API 서버 없으면 직접 OpenAI 호출
                        if result is None:
                            try:
                                result = _analyze_with_openai_direct(raw_url, vid)
                            except Exception as e:
                                err_msg = str(e).lower()
                                if "not found" in err_msg or "404" in err_msg or "video unavailable" in err_msg:
                                    _show_url_error("not_found", raw_url)
                                else:
                                    _show_url_error("error", raw_url)
                                result = None

                        # ── 5분 초과 체크 ──────────────────────────
                        if result is None and (_time.time() - _analysis_start) >= TIMEOUT_SEC:
                            st.session_state.timeout_feedback_sent = False
                            _show_url_error("timeout", raw_url)
                        elif result:
                            st.session_state.timeout_feedback_sent = False
                            st.session_state.latest_result = result
                            st.session_state.video_id = result["video_id"]
                            st.session_state.analyzed_url = raw_url
                            _yt = result.get("video_info") or _fetch_youtube_info(raw_url)
                            st.session_state.analyzed_title    = _yt.get("title", "—")
                            st.session_state.analyzed_views    = _yt.get("view_count")
                            st.session_state.analyzed_duration = _yt.get("duration") or _yt.get("duration_sec")
                            # ── 분석 결과 DB 저장 ──────────────────
                            _save_analysis_to_db(result, raw_url)
                            st.session_state.show_report = True
                            st.rerun()

                    except Exception as e:
                        elapsed = _time.time() - _analysis_start
                        if elapsed >= TIMEOUT_SEC:
                            st.session_state.timeout_feedback_sent = False
                            _show_url_error("timeout", raw_url)
                        else:
                            _show_url_error("error", raw_url)

    elif submitted:
        _, c, _ = st.columns([1, 3, 1])
        with c:
            st.markdown("""
            <div style='text-align:center; color:#856404; background:#fff3cd;
                        border:1px solid #ffc107; border-radius:10px;
                        padding:12px 20px; font-size:0.92rem;'>
              URL을 입력해주세요.
            </div>
            """, unsafe_allow_html=True)

    # 바로가기: 최근 결과가 있으면 리포트 버튼 표시
    if st.session_state.get("latest_result"):
        vid = st.session_state.latest_result.get("video_id", "")
        cat = st.session_state.latest_result.get("analysis_result", {}).get("category", "?")
        meta = CATEGORY_META.get(cat, CATEGORY_META["C5"])
        st.markdown(f"""
        <div style="background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.07);
                    border-radius:12px;padding:14px 20px;margin-top:6px;display:flex;
                    align-items:center;justify-content:space-between;">
          <div style="display:flex;align-items:center;gap:12px;">
            <span style="font-size:1.5rem">{meta['icon']}</span>
            <div>
              <div style="font-size:0.8rem;color:#3d4263">최근 분석 결과</div>
              <div style="font-size:0.9rem;color:#8b90b5">{vid}</div>
            </div>
          </div>
          <span style="font-size:0.78rem;color:#3d4263">{cat} · {meta['label']}</span>
        </div>
        """, unsafe_allow_html=True)

        if st.button("📊 상세 리포트 보기", key="view_report"):
            st.session_state.video_id = vid
            st.session_state.show_report = True
            st.rerun()

    # 시스템 현황은 사이드바로 이동



# ─────────────────────────────────────────
# DB 안전 조회 헬퍼
# ─────────────────────────────────────────

def _generate_excel_export() -> bytes | None:
    """DB에서 분석 결과를 읽어 엑셀 파일(bytes)로 반환"""
    try:
        import io, json
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from database_manager import db_manager
        from database_models import AnalysisResults, Contents

        session = db_manager.get_session()
        try:
            rows = (
                session.query(AnalysisResults)
                .join(Contents, AnalysisResults.video_id == Contents.video_id)
                .order_by(AnalysisResults.created_at.desc())
                .all()
            )
            if not rows:
                return None
        finally:
            session.close()

        # raw_response JSON 파싱 헬퍼
        def _parse_raw(raw):
            try:
                d = json.loads(raw or "")
                return d.get("subtitle_content", ""), d.get("frame_analysis", "")
            except Exception:
                return "", ""

        wb = Workbook()
        ws = wb.active
        ws.title = "분석결과"

        headers = ["Video ID", "분류", "신뢰도", "상태", "판단 근거",
                   "자막 내용", "프레임 분석", "모델", "처리시간(초)", "분석 시각"]
        ws.append(headers)

        # 헤더 스타일
        header_fill = PatternFill("solid", start_color="CC0000", end_color="CC0000")
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        cat_colors  = {"C1": "FECDD3", "C2": "FED7AA", "C3": "FEF08A",
                       "C4": "E9D5FF", "C5": "BBF7D0"}
        stat_colors = {"AUTO_APPROVE": "BBF7D0", "HUMAN_REVIEW": "FEF08A",
                       "AUTO_REJECT": "FECDD3", "ANALYSIS_FAILED": "E5E7EB"}

        for r in rows:
            subtitle, frame = _parse_raw(r.raw_response)
            conf = f"{r.confidence_score:.0%}" if r.confidence_score is not None else "-"
            proc = f"{r.processing_time:.1f}s" if r.processing_time is not None else "-"
            created = r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "-"

            ws.append([
                r.video_id, r.c_category, conf, r.status.value,
                r.reasoning_log or "", subtitle, frame,
                r.model_used or "", proc, created
            ])

            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 80
            cat = r.c_category or ""
            stat = r.status.value if r.status else ""

            for cell in ws[row_idx]:
                cell.font = Font(name="Arial", size=9)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 분류 컬럼 색상
            if cat in cat_colors:
                ws.cell(row_idx, 2).fill = PatternFill("solid", start_color=cat_colors[cat], end_color=cat_colors[cat])
                ws.cell(row_idx, 2).font = Font(name="Arial", size=9, bold=True)
                ws.cell(row_idx, 2).alignment = Alignment(horizontal="center", vertical="top")

            # 상태 컬럼 색상
            if stat in stat_colors:
                ws.cell(row_idx, 4).fill = PatternFill("solid", start_color=stat_colors[stat], end_color=stat_colors[stat])
                ws.cell(row_idx, 4).alignment = Alignment(horizontal="center", vertical="top")

        # 컬럼 너비
        for col, width in zip("ABCDEFGHIJ", [18, 8, 8, 16, 50, 30, 50, 14, 12, 20]):
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    except Exception as e:
        logger.error(f"엑셀 export 실패: {e}")
        return None


def safe_db_counts() -> tuple:
    """DB 카운트를 안전하게 조회. 실패 시 Mock 반환."""
    try:
        from database_manager import db_manager
        from database_models import Contents, AnalysisResults, UserFeedback, ValidationLabels, ReviewStatus
        session = db_manager.get_session()
        try:
            total_c = session.query(Contents).count()
            total_a = session.query(AnalysisResults).count()
            total_f = session.query(UserFeedback).count()
            pending  = session.query(ValidationLabels).filter(
                ValidationLabels.review_status == ReviewStatus.PENDING).count()
            return total_c, total_a, total_f, pending
        finally:
            session.close()
    except Exception:
        return (random.randint(80, 200), random.randint(80, 200),
                random.randint(10, 80),  random.randint(0, 20))


def safe_db_recent_rows() -> list:
    """최근 분석 이력을 안전하게 조회. 실패 시 Mock 반환."""
    try:
        from database_manager import db_manager
        from database_models import AnalysisResults, Contents
        session = db_manager.get_session()
        try:
            rows = session.query(AnalysisResults).join(Contents).order_by(
                AnalysisResults.created_at.desc()).limit(8).all()
            if not rows:
                raise ValueError("empty")
            return [{
                "Video ID": r.video_id,
                "제목": ((r.content.title or "")[:40] + "…") if r.content and r.content.title else "—",
                "카테고리": r.c_category,
                "신뢰도": r.confidence_score,
                "상태": r.status.value,
                "분석 시각": r.created_at.strftime("%m/%d %H:%M"),
            } for r in rows]
        finally:
            session.close()
    except Exception:
        mock_rows = []
        for _ in range(6):
            cat = random.choice(["C1", "C2", "C3", "C4", "C5"])
            mock_rows.append({
                "Video ID": f"mock_{random.randint(10000,99999)}",
                "제목": random.choice(["🔥충격 100만원 비법 공개", "Python 기초 강의",
                                       "TTS 자동생성 영상", "맛집 리뷰 솔직 후기", "무단 복사 의심 영상"]),
                "카테고리": cat,
                "신뢰도": f"{random.uniform(0.55, 0.97):.2f}",
                "상태": CATEGORY_META[cat]["badge"],
                "분석 시각": (datetime.now() - timedelta(minutes=random.randint(1, 180))).strftime("%m/%d %H:%M"),
            })
        return mock_rows


def safe_db_category_dist() -> dict:
    """카테고리 분포를 안전하게 조회. 실패 시 Mock 반환."""
    try:
        from database_manager import db_manager
        from database_models import AnalysisResults
        session = db_manager.get_session()
        try:
            cat_counts = {"C1": 0, "C2": 0, "C3": 0, "C4": 0, "C5": 0}
            results = session.query(AnalysisResults).order_by(
                AnalysisResults.created_at.desc()).limit(100).all()
            for r in results:
                if r.c_category in cat_counts:
                    cat_counts[r.c_category] += 1
            if sum(cat_counts.values()) == 0:
                raise ValueError("empty")
            return cat_counts
        finally:
            session.close()
    except Exception:
        return {k: random.randint(2, 25) for k in ["C1", "C2", "C3", "C4", "C5"]}




def safe_db_feedback_action_dist() -> dict:
    """user_action별 피드백 건수 조회. 실패 시 Mock 반환."""
    try:
        from database_manager import db_manager
        from database_models import UserFeedback
        from sqlalchemy import func as _func
        session = db_manager.get_session()
        try:
            rows = session.query(
                UserFeedback.user_action,
                _func.count(UserFeedback.id).label("cnt")
            ).group_by(UserFeedback.user_action).all()
            result = {r.user_action: r.cnt for r in rows}
            if not result:
                raise ValueError("empty")
            return result
        finally:
            session.close()
    except Exception:
        return {
            "like":          random.randint(15, 60),
            "dislike":       random.randint(5, 30),
            "block_channel": random.randint(3, 20),
            "report":        random.randint(2, 15),
            "opinion":       random.randint(1, 10),
        }


def _build_yt_url(video_id: str, url: str = "") -> str:
    """video_id 또는 저장된 url로 YouTube 링크 생성."""
    if url and url.startswith("http"):
        return url
    if video_id:
        import re
        if re.match(r'^[A-Za-z0-9_-]{11}$', video_id):
            return f"https://www.youtube.com/shorts/{video_id}"
    return ""


def safe_db_feedback_rows(limit: int = 200, action_filter: str = "전체") -> list:
    """최근 피드백 목록 조회 (Contents JOIN으로 실제 YouTube URL 포함). 실패 시 Mock 반환."""
    try:
        from database_manager import db_manager
        from database_models import UserFeedback, Contents
        session = db_manager.get_session()
        try:
            q = (session.query(UserFeedback, Contents.url, Contents.title)
                 .outerjoin(Contents, UserFeedback.video_id == Contents.video_id)
                 .order_by(UserFeedback.created_at.desc()))
            if action_filter != "전체":
                q = q.filter(UserFeedback.user_action == action_filter)
            rows = q.limit(limit).all()
            if not rows:
                raise ValueError("empty")
            result = []
            for r, url, title in rows:
                # feedback_context에 source_url 있으면 우선 사용 (직접 저장 경로)
                ctx = r.feedback_context or {}
                source_url = ctx.get("source_url", "") if isinstance(ctx, dict) else ""
                yt_url = source_url or _build_yt_url(r.video_id, url or "")
                result.append({
                    "feedback_id":   r.feedback_id,
                    "video_id":      r.video_id,
                    "video_title":   title or "",
                    "youtube_url":   yt_url,
                    "user_action":   r.user_action,
                    "feedback_text": r.feedback_text or "",
                    "ip_address":    r.ip_address or "—",
                    "is_processed":  r.is_processed,
                    "created_at":    r.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                })
            return result
        finally:
            session.close()
    except Exception:
        _actions = ["like", "dislike", "block_channel", "report", "opinion"]
        _yt_ids  = ["dQw4w9WgXcQ", "9bZkp7q19f0", "kJQP7kiw5Fk", "JGwWNGJdvx8", "RgKAFK5djSk"]
        _texts   = ["[AI분류:C1(어그로/스팸)→사용자제안:C5(정상영상)] 정상 영상 같습니다.", "", "", "명백한 어그로입니다.", ""]
        mock = []
        for i in range(min(limit, 30)):
            act = random.choice(_actions) if action_filter == "전체" else action_filter
            vid = random.choice(_yt_ids)
            mock.append({
                "feedback_id":   f"feedback_{random.randint(10000000,99999999):08x}",
                "video_id":      vid,
                "video_title":   "",
                "youtube_url":   f"https://www.youtube.com/shorts/{vid}",
                "user_action":   act,
                "feedback_text": random.choice(_texts),
                "ip_address":    f"192.168.{random.randint(0,255)}.{random.randint(1,254)}",
                "is_processed":  random.choice([True, False]),
                "created_at":    (datetime.now() - timedelta(minutes=random.randint(1, 2880))).strftime("%Y-%m-%d %H:%M:%S"),
            })
        return mock


def safe_db_feedback_trend(days: int = 7) -> list:
    """일별 피드백 건수 트렌드. 실패 시 Mock 반환."""
    try:
        from database_manager import db_manager
        from database_models import UserFeedback
        from sqlalchemy import func as _func, cast, Date
        session = db_manager.get_session()
        try:
            cutoff = datetime.now() - timedelta(days=days)
            rows = session.query(
                cast(UserFeedback.created_at, Date).label("day"),
                UserFeedback.user_action,
                _func.count(UserFeedback.id).label("cnt")
            ).filter(UserFeedback.created_at >= cutoff)             .group_by("day", UserFeedback.user_action)             .order_by("day").all()
            if not rows:
                raise ValueError("empty")
            return [{"day": str(r.day), "action": r.user_action, "cnt": r.cnt} for r in rows]
        finally:
            session.close()
    except Exception:
        trend = []
        for d in range(days):
            day = (datetime.now() - timedelta(days=days - 1 - d)).strftime("%Y-%m-%d")
            for act in ["like", "dislike", "block_channel", "report", "opinion"]:
                trend.append({"day": day, "action": act, "cnt": random.randint(0, 12)})
        return trend


# ─────────────────────────────────────────
# 관리자 DB 액션 헬퍼
# ─────────────────────────────────────────

def admin_mark_processed(feedback_id: str) -> bool:
    """① 피드백 처리 완료 체크 (is_processed=True, processed_at=now)"""
    try:
        from database_manager import db_manager
        from database_models import UserFeedback
        session = db_manager.get_session()
        try:
            fb = session.query(UserFeedback).filter(UserFeedback.feedback_id == feedback_id).first()
            if fb:
                fb.is_processed = True
                fb.processed_at = datetime.now()
                session.commit()
                return True
        finally:
            session.close()
    except Exception as e:
        logger.error(f"admin_mark_processed 오류: {e}")
    return False


def admin_promote_to_hitl(video_id: str, result_id: str = None) -> bool:
    """② 피드백 → HITL 큐 승격 (ValidationLabels에 is_hard_example=True로 등록)"""
    try:
        from database_manager import db_manager
        from database_models import ValidationLabels, ReviewStatus
        import uuid as _uuid
        session = db_manager.get_session()
        try:
            # 이미 존재하는 ValidationLabel 확인
            existing = session.query(ValidationLabels).filter(
                ValidationLabels.video_id == video_id
            ).first()
            if existing:
                existing.is_hard_example = True
                existing.review_status = ReviewStatus.PENDING
            else:
                label = ValidationLabels(
                    label_id=f"label_{_uuid.uuid4().hex[:12]}",
                    video_id=video_id,
                    result_id=result_id,
                    is_hard_example=True,
                    review_status=ReviewStatus.PENDING,
                )
                session.add(label)
            session.commit()
            return True
        finally:
            session.close()
    except Exception as e:
        logger.error(f"admin_promote_to_hitl 오류: {e}")
    return False


def admin_set_ground_truth(video_id: str, category: str, comment: str = "") -> bool:
    """③ Ground Truth 카테고리 지정 (ValidationLabels.ground_truth_category 업데이트)"""
    try:
        from database_manager import db_manager
        from database_models import ValidationLabels, ReviewStatus
        import uuid as _uuid
        session = db_manager.get_session()
        try:
            label = session.query(ValidationLabels).filter(
                ValidationLabels.video_id == video_id
            ).first()
            if not label:
                label = ValidationLabels(
                    label_id=f"label_{_uuid.uuid4().hex[:12]}",
                    video_id=video_id,
                    is_hard_example=True,
                )
                session.add(label)
            label.ground_truth_category = category
            label.review_status = ReviewStatus.APPROVED
            label.human_reviewer_id = "admin"
            label.reviewed_at = datetime.now()
            if comment:
                label.review_comments = comment
            session.commit()
            return True
        finally:
            session.close()
    except Exception as e:
        logger.error(f"admin_set_ground_truth 오류: {e}")
    return False


def admin_save_comment(feedback_id: str, comment: str) -> bool:
    """④ 피드백에 관리자 메모 저장 (feedback_context JSON에 admin_comment 필드)"""
    try:
        from database_manager import db_manager
        from database_models import UserFeedback
        session = db_manager.get_session()
        try:
            fb = session.query(UserFeedback).filter(UserFeedback.feedback_id == feedback_id).first()
            if fb:
                ctx = fb.feedback_context or {}
                ctx["admin_comment"] = comment
                ctx["admin_comment_at"] = datetime.now().isoformat()
                fb.feedback_context = ctx
                session.commit()
                return True
        finally:
            session.close()
    except Exception as e:
        logger.error(f"admin_save_comment 오류: {e}")
    return False


def show_admin_feedback():
    """관리자용 피드백 현황 전체 화면"""

    ACTION_LABELS = {
        "like":          "👍 좋아요",
        "dislike":       "👎 싫어요",
        "block_channel": "🚫 채널 차단",
        "report":        "📢 신고하기",
        "opinion":       "💬 의견 보내기",
    }
    ACTION_COLORS = {
        "like":          "#10b981",
        "dislike":       "#f59e0b",
        "block_channel": "#8b5cf6",
        "report":        "#ef4444",
        "opinion":       "#3b82f6",
    }

    st.markdown("<div style='margin-top:-18vh'></div>", unsafe_allow_html=True)

    st.markdown("""
    <div style="margin-bottom:28px">
      <div style="font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;
                  color:#1a1a2e;letter-spacing:-0.5px">🛠️ 관리자 피드백 현황</div>
      <div style="font-size:0.85rem;color:#888;margin-top:4px">
        user_feedback 테이블 · 실시간 데이터
      </div>
    </div>
    """, unsafe_allow_html=True)

    # KPI 카드
    action_dist = safe_db_feedback_action_dist()
    total_fb = sum(action_dist.values())
    likes    = action_dist.get("like", 0)
    reports  = action_dist.get("report", 0) + action_dist.get("block_channel", 0)
    opinions = action_dist.get("opinion", 0)

    kpi_cols = st.columns(4)
    for col, (label, val, fg, bg) in zip(kpi_cols, [
        ("📊 전체 피드백",  total_fb, "#7c83fd", "#f0f0ff"),
        ("👍 좋아요",       likes,    "#10b981", "#f0fdf4"),
        ("📢 신고 / 차단",  reports,  "#ef4444", "#fff5f5"),
        ("💬 의견 제출",    opinions, "#3b82f6", "#eff6ff"),
    ]):
        with col:
            st.markdown(f"""
            <div style="background:{bg};border:1px solid {fg}22;border-radius:14px;
                        padding:20px 22px;text-align:center;">
              <div style="font-size:1.8rem;font-weight:800;color:{fg};
                          font-family:'Syne',sans-serif;line-height:1">{val:,}</div>
              <div style="font-size:0.78rem;color:#888;margin-top:6px;
                          text-transform:uppercase;letter-spacing:0.08em">{label}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)

    # 차트
    col_donut, col_trend = st.columns([1, 2])

    with col_donut:
        st.markdown("<div style='font-family:Syne,sans-serif;font-size:0.7rem;font-weight:700;"
                    "letter-spacing:0.12em;text-transform:uppercase;color:#aaa;"
                    "margin-bottom:12px'>액션 유형 분포</div>", unsafe_allow_html=True)
        labels_d = [ACTION_LABELS.get(k, k) for k in action_dist]
        values_d = list(action_dist.values())
        colors_d = [ACTION_COLORS.get(k, "#999") for k in action_dist]
        fig_donut = go.Figure(go.Pie(
            labels=labels_d, values=values_d, hole=0.58,
            marker=dict(colors=colors_d, line=dict(color="#ffffff", width=2)),
            hovertemplate="%{label}: %{value}건 (%{percent})<extra></extra>",
            textinfo="none",
        ))
        fig_donut.add_annotation(
            text=f"<b>{total_fb}</b><br><span style='font-size:10px'>총 피드백</span>",
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=18, color="#333333", family="Syne"),
            xref="paper", yref="paper"
        )
        fig_donut.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0, r=0, t=0, b=0), height=260, showlegend=True,
            legend=dict(orientation="v", x=1.02, y=0.5,
                        font=dict(size=11, color="#555"), bgcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_donut, use_container_width=True, config={"displayModeBar": False})

    with col_trend:
        st.markdown("<div style='font-family:Syne,sans-serif;font-size:0.7rem;font-weight:700;"
                    "letter-spacing:0.12em;text-transform:uppercase;color:#aaa;"
                    "margin-bottom:12px'>일별 피드백 트렌드 (최근 7일)</div>", unsafe_allow_html=True)
        trend_data = safe_db_feedback_trend(days=7)
        df_trend = pd.DataFrame(trend_data)
        fig_trend = go.Figure()
        for act, color in ACTION_COLORS.items():
            sub = df_trend[df_trend["action"] == act] if not df_trend.empty else pd.DataFrame()
            if sub.empty:
                continue
            fig_trend.add_trace(go.Scatter(
                x=sub["day"], y=sub["cnt"], mode="lines+markers",
                name=ACTION_LABELS.get(act, act),
                line=dict(color=color, width=2), marker=dict(size=6, color=color),
                hovertemplate=f"{ACTION_LABELS.get(act, act)}: %{{y}}건<extra></extra>",
            ))
        fig_trend.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(showgrid=False, tickfont=dict(size=10, color="#999")),
            yaxis=dict(gridcolor="rgba(0,0,0,0.07)", tickfont=dict(size=10, color="#999"), zeroline=False),
            legend=dict(orientation="h", y=-0.22, font=dict(size=10, color="#666"), bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=0, r=0, t=0, b=40), height=260, hovermode="x unified",
        )
        st.plotly_chart(fig_trend, use_container_width=True, config={"displayModeBar": False})

    st.markdown("<hr style='border-color:#f0f0f0;margin:24px 0'>", unsafe_allow_html=True)

    # 필터 + 테이블
    st.markdown("<div style='font-family:Syne,sans-serif;font-size:0.7rem;font-weight:700;"
                "letter-spacing:0.12em;text-transform:uppercase;color:#aaa;"
                "margin-bottom:14px'>피드백 상세 목록</div>", unsafe_allow_html=True)

    filter_cols = st.columns([2, 2, 1])
    with filter_cols[0]:
        action_options = ["전체", "like", "dislike", "block_channel", "report", "opinion"]
        action_labels  = ["전체", "👍 좋아요", "👎 싫어요", "🚫 채널 차단", "📢 신고하기", "💬 의견 보내기"]
        selected_label = st.selectbox("액션 유형 필터", options=action_labels, index=0, key="admin_action_filter")
        selected_action = action_options[action_labels.index(selected_label)]
    with filter_cols[1]:
        search_vid = st.text_input("Video ID 검색", placeholder="video_12345", key="admin_vid_search")
    with filter_cols[2]:
        limit_n = st.selectbox("표시 건수", [50, 100, 200], index=0, key="admin_limit")

    rows = safe_db_feedback_rows(limit=limit_n, action_filter=selected_action)
    if search_vid.strip():
        rows = [r for r in rows if search_vid.strip().lower() in r["video_id"].lower()]

    if not rows:
        st.info("조건에 맞는 피드백이 없습니다.")
    else:
        processed_cnt = sum(1 for r in rows if r["is_processed"])
        text_cnt      = sum(1 for r in rows if r["feedback_text"])
        pending_cnt   = len(rows) - processed_cnt

        sum_cols = st.columns(3)
        for col, (lbl, val, color) in zip(sum_cols, [
            (f"조회된 피드백 {len(rows)}건", f"처리 완료 {processed_cnt}건 / 대기 {pending_cnt}건", "#3b82f6"),
            ("의견 텍스트 포함",             f"{text_cnt}건",                                        "#f59e0b"),
            ("미처리 (is_processed=False)",  f"{pending_cnt}건",                                     "#ef4444"),
        ]):
            with col:
                st.markdown(f"""
                <div style="background:#f8f9fa;border-left:3px solid {color};
                            border-radius:8px;padding:10px 14px;margin-bottom:12px">
                  <div style="font-size:0.75rem;color:#888">{lbl}</div>
                  <div style="font-size:1rem;font-weight:700;color:{color};margin-top:2px">{val}</div>
                </div>
                """, unsafe_allow_html=True)

        df_rows = []
        for r in rows:
            vid    = r["video_id"]
            yt_url = r.get("youtube_url", "") or _build_yt_url(vid)
            df_rows.append({
                "시각":      r["created_at"],
                "영상 링크": yt_url,
                "Video ID":  vid,
                "액션":      ACTION_LABELS.get(r["user_action"], r["user_action"]),
                "의견 내용": r["feedback_text"] if r["feedback_text"] else "—",
                "처리":      "✅ 완료" if r["is_processed"] else "⏳ 대기",
                "IP":        r["ip_address"],
                "ID":        r["feedback_id"],
            })
        df = pd.DataFrame(df_rows)

        st.dataframe(
            df, use_container_width=True, hide_index=True,
            column_config={
                "시각":      st.column_config.TextColumn("🕒 시각",      width=140),
                "영상 링크": st.column_config.LinkColumn(
                    "🔗 영상 바로가기",
                    display_text="▶ YouTube",
                    width=120,
                ),
                "Video ID":  st.column_config.TextColumn("🎬 Video ID",  width=130),
                "액션":      st.column_config.TextColumn("⚡ 액션",      width=130),
                "의견 내용": st.column_config.TextColumn("💬 의견 내용", width=250),
                "처리":      st.column_config.TextColumn("✅ 처리",      width=80),
                "IP":        st.column_config.TextColumn("🌐 IP",        width=120),
                "ID":        st.column_config.TextColumn("🔑 피드백 ID", width=160),
            },
            height=420,
        )

        col_csv, col_spacer = st.columns([1, 3])
        with col_csv:
            csv_data = df.to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label="⬇️ CSV 다운로드",
                data=csv_data,
                file_name=f"feedback_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv",
                key="admin_csv_dl"
            )

        # ── 관리자 액션 패널 ────────────────────────────────────────
        st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
        st.markdown("""
        <div style="font-family:'Syne',sans-serif;font-size:0.7rem;font-weight:700;
                    letter-spacing:0.12em;text-transform:uppercase;color:#aaa;
                    margin-bottom:16px">⚡ 관리자 액션</div>
        """, unsafe_allow_html=True)

        # 피드백 선택 드롭다운
        feedback_options = {
            f"{ACTION_LABELS.get(r['user_action'], r['user_action'])}  |  {r['video_id']}  |  {r['created_at']}": r
            for r in rows
        }
        selected_key = st.selectbox(
            "처리할 피드백 선택",
            options=list(feedback_options.keys()),
            index=0,
            key="admin_selected_feedback",
            help="위 테이블에서 처리할 피드백을 선택하세요"
        )
        sel = feedback_options[selected_key]

        # 선택된 피드백 요약 카드
        action_color = ACTION_COLORS.get(sel["user_action"], "#888")
        action_label = ACTION_LABELS.get(sel["user_action"], sel["user_action"])
        yt_url_sel   = sel.get("youtube_url", "") or _build_yt_url(sel["video_id"])
        yt_link      = f'<a href="{yt_url_sel}" target="_blank" style="color:{action_color};text-decoration:none">▶ YouTube에서 보기 →</a>' if yt_url_sel else ""
        existing_comment = (sel.get("feedback_context") or {}).get("admin_comment", "") if isinstance(sel.get("feedback_context"), dict) else ""

        st.markdown(f"""
        <div style="background:#f8f9fa;border:1px solid #e8e8f0;border-left:4px solid {action_color};
                    border-radius:10px;padding:16px 20px;margin-bottom:20px">
          <div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap">
            <span style="background:{action_color}18;color:{action_color};font-size:0.8rem;
                         font-weight:700;padding:3px 10px;border-radius:20px">{action_label}</span>
            <span style="font-family:monospace;font-size:0.82rem;color:#555">{sel['video_id']}</span>
            <span style="font-size:0.78rem;color:#aaa">{sel['created_at']}</span>
            {"<span style='font-size:0.78rem;background:#10b98118;color:#10b981;padding:2px 8px;border-radius:10px'>✅ 처리완료</span>" if sel['is_processed'] else "<span style='font-size:0.78rem;background:#f59e0b18;color:#f59e0b;padding:2px 8px;border-radius:10px'>⏳ 미처리</span>"}
            {yt_link}
          </div>
          {f'<div style="margin-top:10px;font-size:0.83rem;color:#555;background:#fff;border-radius:6px;padding:8px 12px">💬 {sel["feedback_text"]}</div>' if sel.get("feedback_text") else ""}
        </div>
        """, unsafe_allow_html=True)

        # 4개 액션 탭
        tab1, tab2, tab3, tab4 = st.tabs([
            "① ✅ 처리 완료",
            "② 🔺 HITL 큐 승격",
            "③ 🏷️ Ground Truth",
            "④ 📝 메모 남기기",
        ])

        # ── ① 처리 완료 체크 ─────────────────────────
        with tab1:
            st.markdown("""
            <div style="font-size:0.85rem;color:#555;margin:8px 0 16px 0">
            피드백을 <b>처리 완료</b>로 표시합니다. 
            <code>is_processed=True</code>, <code>processed_at=now()</code>로 업데이트됩니다.
            </div>
            """, unsafe_allow_html=True)

            if sel["is_processed"]:
                st.success("이미 처리 완료된 피드백입니다.")
            else:
                if st.button("✅ 처리 완료로 표시", key="btn_mark_processed", type="primary"):
                    ok = admin_mark_processed(sel["feedback_id"])
                    if ok:
                        st.success(f"✅ `{sel['feedback_id']}` 처리 완료로 업데이트했습니다.")
                        st.rerun()
                    else:
                        st.error("DB 업데이트 실패 (DB 연결 없이 Mock 데이터 상태입니다)")

        # ── ② HITL 큐 승격 ───────────────────────────
        with tab2:
            st.markdown("""
            <div style="font-size:0.85rem;color:#555;margin:8px 0 16px 0">
            이 영상을 <b>Hard Example</b>로 등록하여 HITL 검토 큐에 추가합니다.<br>
            <code>ValidationLabels</code> 테이블에 <code>is_hard_example=True</code>로 기록되며,
            관리자의 Ground Truth 검토 대상이 됩니다.
            </div>
            """, unsafe_allow_html=True)

            hitl_col1, hitl_col2 = st.columns([3, 1])
            with hitl_col1:
                hitl_reason = st.text_input(
                    "승격 사유 (선택)",
                    placeholder="예: 사용자 신고 다수, AI 판단 불확실 등",
                    key="hitl_reason_input"
                )
            with hitl_col2:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("🔺 HITL 큐에 추가", key="btn_hitl_promote", type="primary"):
                    ok = admin_promote_to_hitl(sel["video_id"])
                    if ok:
                        # 동시에 처리 완료도 표시
                        admin_mark_processed(sel["feedback_id"])
                        st.success(f"✅ `{sel['video_id']}` 를 HITL Hard Example 큐에 추가했습니다.")
                        st.rerun()
                    else:
                        st.error("DB 업데이트 실패 (DB 연결 없이 Mock 데이터 상태입니다)")

            st.markdown("""
            <div style="background:#f0f4ff;border-radius:8px;padding:10px 14px;
                        font-size:0.78rem;color:#555;margin-top:8px">
            💡 <b>HITL 큐란?</b> AI가 모호하게 판단한 영상을 사람이 직접 검토해
            정확한 카테고리를 매기는 파이프라인입니다. 여기서 쌓인 데이터가
            Qwen 모델의 재학습에 사용됩니다.
            </div>
            """, unsafe_allow_html=True)

        # ── ③ Ground Truth 카테고리 지정 ─────────────
        with tab3:
            st.markdown("""
            <div style="font-size:0.85rem;color:#555;margin:8px 0 16px 0">
            관리자가 직접 <b>정답 카테고리(Ground Truth)</b>를 지정합니다.
            <code>ValidationLabels.ground_truth_category</code>에 저장되어 모델 학습 데이터로 활용됩니다.
            </div>
            """, unsafe_allow_html=True)

            CATEGORY_OPTIONS = {
                "C1 — 어그로 / 스팸":       "C1",
                "C2 — 가짜뉴스 / 허위정보": "C2",
                "C3 — 혐오 / 폭력":         "C3",
                "C4 — 저품질 / 광고":       "C4",
                "C5 — 정상 영상":           "C5",
            }
            gt_col1, gt_col2 = st.columns([2, 2])
            with gt_col1:
                gt_label = st.selectbox(
                    "정답 카테고리 선택",
                    options=list(CATEGORY_OPTIONS.keys()),
                    key="gt_category_select"
                )
                gt_category = CATEGORY_OPTIONS[gt_label]
            with gt_col2:
                gt_comment = st.text_input(
                    "검토 코멘트 (선택)",
                    placeholder="예: 영상 내 허위 주장 확인됨",
                    key="gt_comment_input"
                )

            if st.button(f"🏷️ {gt_label} 으로 Ground Truth 저장", key="btn_set_gt", type="primary"):
                ok = admin_set_ground_truth(sel["video_id"], gt_category, gt_comment)
                if ok:
                    admin_mark_processed(sel["feedback_id"])
                    st.success(f"✅ `{sel['video_id']}` → **{gt_label}** 으로 Ground Truth 지정 완료!")
                    st.markdown(f"""
                    <div style="background:#f0fdf4;border:1px solid #10b98144;border-radius:8px;
                                padding:10px 14px;font-size:0.82rem;color:#059669;margin-top:8px">
                    🎓 이 데이터는 다음 지식증류 배치에 포함되어 Qwen 모델 학습에 사용됩니다.
                    </div>
                    """, unsafe_allow_html=True)
                    st.rerun()
                else:
                    st.error("DB 업데이트 실패 (DB 연결 없이 Mock 데이터 상태입니다)")

        # ── ④ 메모 남기기 ─────────────────────────────
        with tab4:
            st.markdown("""
            <div style="font-size:0.85rem;color:#555;margin:8px 0 16px 0">
            피드백에 <b>관리자 메모</b>를 남깁니다.
            <code>UserFeedback.feedback_context</code> JSON의 <code>admin_comment</code> 필드에 저장됩니다.
            </div>
            """, unsafe_allow_html=True)

            if existing_comment:
                st.markdown(f"""
                <div style="background:#fffbeb;border:1px solid #f59e0b44;border-radius:8px;
                            padding:10px 14px;font-size:0.83rem;color:#92400e;margin-bottom:12px">
                📌 기존 메모: {existing_comment}
                </div>
                """, unsafe_allow_html=True)

            new_comment = st.text_area(
                "메모 내용",
                value=existing_comment,
                placeholder="예: 동일 IP에서 반복 신고 확인, 채널 모니터링 필요 등",
                height=100,
                key="admin_memo_input"
            )
            if st.button("💾 메모 저장", key="btn_save_memo", type="primary"):
                if new_comment.strip():
                    ok = admin_save_comment(sel["feedback_id"], new_comment.strip())
                    if ok:
                        st.success("✅ 메모가 저장되었습니다.")
                        st.rerun()
                    else:
                        st.error("DB 업데이트 실패 (DB 연결 없이 Mock 데이터 상태입니다)")
                else:
                    st.warning("메모 내용을 입력해주세요.")

# ─────────────────────────────────────────
# 메인 진입점
# ─────────────────────────────────────────

def render_sidebar():
    """사이드바 렌더링 - 반드시 main() 안에서 호출"""
    with st.sidebar:
        st.markdown("""
        <div style="padding:8px 0 20px 0">
          <p style="font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;
                    color:#cc0000;margin:0;letter-spacing:-0.5px">Shorts</p>
          <p style="font-family:'Syne',sans-serif;font-size:1.15rem;font-weight:700;
                    color:#333333;margin:0;letter-spacing:-0.5px">Check</p>
          <p style="font-size:0.72rem;color:#888888;margin-top:4px;letter-spacing:0.06em">
            이 영상에 대해 어떻게 생각하시나요?</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("---")

        # API 상태
        api_ok = check_api_health()
        chip_class = "chip-online" if api_ok else "chip-offline"
        chip_text  = "● API 서버 정상" if api_ok else "● API 서버 오프라인 (Mock)"
        st.markdown(
            f"<span class='status-chip {chip_class}'>{chip_text}</span>",
            unsafe_allow_html=True
        )

        # ── 아코디언 1: 분류 체계 ──
        with st.expander("📋 분류 체계", expanded=False):
            for k, v in CATEGORY_META.items():
                color = {"danger": "#dc2626", "warning": "#d97706", "safe": "#059669"}[v["verdict"]]
                st.markdown(
                    f"<div style='display:flex;align-items:center;gap:8px;padding:6px 0;"
                    f"border-bottom:1px solid #f0f0f0;font-size:0.84rem;'>"
                    f"<span style='font-size:1rem'>{v['icon']}</span>"
                    f"<b style='color:{color}'>{k}</b>"
                    f"<span style='color:#888;font-size:0.8rem'>{v['label']}</span></div>",
                    unsafe_allow_html=True
                )

        # ── 아코디언 2: 카테고리 분포 ──
        with st.expander("📊 카테고리 분포", expanded=False):
            cat_counts = safe_db_category_dist()
            total = sum(cat_counts.values()) or 1
            colors = {"C1": "#ef4444", "C2": "#f97316", "C3": "#f59e0b",
                      "C4": "#8b5cf6", "C5": "#10b981"}
            for k, cnt in cat_counts.items():
                pct = cnt / total * 100
                st.markdown(
                    f"<div style='margin-bottom:8px'>"
                    f"<div style='display:flex;justify-content:space-between;"
                    f"font-size:0.8rem;margin-bottom:3px'>"
                    f"<span style='color:#333;font-weight:600'>{CATEGORY_META[k]['icon']} {k}</span>"
                    f"<span style='color:#888'>{cnt}건 ({pct:.0f}%)</span></div>"
                    f"<div style='background:#f0f0f0;border-radius:4px;height:6px'>"
                    f"<div style='background:{colors[k]};width:{pct:.0f}%;"
                    f"height:6px;border-radius:4px;transition:width 0.3s'></div>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )

        # ── 아코디언 3: 최근 분석 이력 ──
        with st.expander("🕒 최근 분석 이력", expanded=False):
            rows = safe_db_recent_rows()

            CAT_STYLE = {
                "C1": ("#fef2f2", "#dc2626", "어그로"),
                "C2": ("#fff7ed", "#ea580c", "공장형"),
                "C3": ("#fefce8", "#ca8a04", "품질불량"),
                "C4": ("#f5f3ff", "#7c3aed", "무단도용"),
                "C5": ("#f0fdf4", "#16a34a", "정상"),
            }

            for row in rows:
                cat_key = row.get("카테고리", "C5")
                bg, fg, cat_label = CAT_STYLE.get(cat_key, ("#f9f9f9", "#999", cat_key))
                conf = float(row.get("신뢰도", 0))
                conf_color = "#16a34a" if conf >= 0.8 else "#d97706" if conf >= 0.6 else "#dc2626"
                conf_bar = int(conf * 100)

                title = row.get("제목", "—")
                title_short = (title[:18] + "…") if len(title) > 18 else title

                st.markdown(f"""
                <div style="
                    background:#ffffff;
                    border:1px solid #f0f0f0;
                    border-left:3px solid {fg};
                    border-radius:8px;
                    padding:10px 12px;
                    margin-bottom:7px;
                ">
                  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px">
                    <span style="
                        background:{bg};color:{fg};
                        font-size:0.7rem;font-weight:700;
                        padding:2px 8px;border-radius:10px;
                        letter-spacing:0.03em
                    ">{cat_key} {cat_label}</span>
                    <span style="font-size:0.7rem;color:#bbb">{row.get('분석 시각','—')}</span>
                  </div>
                  <div style="font-size:0.82rem;color:#333;font-weight:500;margin-bottom:6px;
                              white-space:nowrap;overflow:hidden;text-overflow:ellipsis">
                    {title_short}
                  </div>
                  <div style="display:flex;align-items:center;gap:8px">
                    <div style="flex:1;height:4px;background:#f0f0f0;border-radius:2px;overflow:hidden">
                      <div style="width:{conf_bar}%;height:100%;background:{conf_color};border-radius:2px"></div>
                    </div>
                    <span style="font-size:0.72rem;font-weight:700;color:{conf_color};min-width:32px;text-align:right">
                      {conf:.0%}
                    </span>
                  </div>
                </div>
                """, unsafe_allow_html=True)

        # ── 시스템 현황 ──
        with st.expander("🖥️ 시스템 현황", expanded=False):
            total_c, total_a, total_f, pending = safe_db_counts()
            for val, lbl in [
                (total_c, "분석된 콘텐츠"),
                (total_a, "분석 결과"),
                (total_f, "사용자 피드백"),
                (pending, "검토 대기 (HITL)"),
            ]:
                st.markdown(
                    f"<div style='display:flex;align-items:center;padding:7px 0;"
                    f"border-bottom:1px solid #f0f0f0;font-size:0.84rem;'>"
                    f"<span style='color:#555;flex:1'>{lbl}</span>"
                    f"<b style='color:#cc0000'>{val:,}</b></div>",
                    unsafe_allow_html=True
                )

        st.markdown("---")
        st.markdown("<div class='section-header'>설정</div>", unsafe_allow_html=True)
        #auto_refresh = st.checkbox("자동 새로고침 (30s)", value=False)
        #if auto_refresh:
        #    import time as _time
        #    _time.sleep(30)
        #    st.rerun()

        if st.button("새로고침", use_container_width=True):
            st.rerun()

        st.markdown("---")
        st.markdown("<div class='section-header'>데이터 내보내기</div>", unsafe_allow_html=True)

        excel_data = _generate_excel_export()
        if excel_data:
            st.download_button(
                label="📥 엑셀로 내보내기",
                data=excel_data,
                file_name=f"shorts_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.caption("분석 데이터가 없습니다.")


def main():
    # ── 세션 초기화 ───────────────────────────
    if "show_report" not in st.session_state:
        st.session_state.show_report = False
    if "video_id" not in st.session_state:
        st.session_state.video_id = None
    if "analyzed_url" not in st.session_state:
        st.session_state.analyzed_url = "—"
    if "analyzed_title" not in st.session_state:
        st.session_state.analyzed_title = "—"
    if "analyzed_views" not in st.session_state:
        st.session_state.analyzed_views = None
    if "analyzed_duration" not in st.session_state:
        st.session_state.analyzed_duration = None

    # ── 사이드바 (main 안에서 호출) ───────────
    render_sidebar()
    _inject_favicon(_LOGO_B64)

    # ── 헤더 위 여백: 시각적 세로 중앙 ──────────
    st.markdown("<div style='height: 18vh'></div>", unsafe_allow_html=True)

    # ── 헤더 ──────────────────────────────────
    st.markdown("""
    <div class="main-header-wrap">
      <h1 class="main-title"><span class="title-red">Shorts</span> Check</h1>
    </div>
    """, unsafe_allow_html=True)

    # ── URL 파라미터 확인 ─────────────────────
    try:
        vid_from_url = st.query_params.get("video_id")
        is_admin     = st.query_params.get("admin") == "true"
    except Exception:
        vid_from_url = None
        is_admin     = False

    # ── 관리자 모드 ───────────────────────────
    if is_admin:
        show_admin_feedback()
        return

    # ── 화면 분기 ─────────────────────────────
    video_id = st.session_state.video_id or vid_from_url
    if (st.session_state.show_report or vid_from_url) and video_id:
        show_report(video_id)
    else:
        show_main()


main()
