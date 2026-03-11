"""
user_feedback_db.py
────────────────────────────────────────────────────────────────
User Feedback 전용 DB 모듈
- 기존 youtube_shorts_detector.db 와 완전히 분리된 별도 파일로 관리
- 관리자 전용 내보내기(CSV / JSON / Excel) 기능 포함
────────────────────────────────────────────────────────────────
"""

import sqlite3
import csv
import json
import os
from datetime import datetime

# ── 경로 설정 ────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
DB_PATH    = os.path.join(BASE_DIR, "user_feedback.db")
EXPORT_DIR = os.path.join(BASE_DIR, "exports")


# ════════════════════════════════════════════════════════════
# 1. DB 초기화
# ════════════════════════════════════════════════════════════
def init_db():
    """
    user_feedback.db 와 테이블을 생성합니다.
    이미 존재하면 아무 작업도 하지 않습니다.
    """
    os.makedirs(EXPORT_DIR, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_feedback (
            feedback_id   INTEGER PRIMARY KEY AUTOINCREMENT,
            video_id      VARCHAR(255) NOT NULL,
            result_id     VARCHAR(255),          -- analysis_results.result_id 참조 (선택)
            feedback_text TEXT        NOT NULL,  -- 사용자가 입력한 의견 내용
            source        VARCHAR(50) DEFAULT 'chrome_extension',  -- 피드백 유입 경로
            created_at    DATETIME    DEFAULT (datetime('now','localtime'))
        )
    """)

    conn.commit()
    conn.close()
    print(f"[DB] user_feedback.db 초기화 완료 → {DB_PATH}")


# ════════════════════════════════════════════════════════════
# 2. 피드백 저장
# ════════════════════════════════════════════════════════════
def save_feedback(video_id: str,
                  feedback_text: str,
                  result_id: str = None,
                  source: str = "chrome_extension") -> int:
    """
    사용자 피드백 1건을 저장합니다.

    Parameters
    ----------
    video_id      : 유튜브 영상 ID
    feedback_text : 사용자가 입력한 의견 텍스트
    result_id     : 연결된 분석 결과 ID (없으면 None)
    source        : 피드백 유입 경로 (기본값: chrome_extension)

    Returns
    -------
    생성된 feedback_id (int)
    """
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    cur.execute("""
        INSERT INTO user_feedback (video_id, result_id, feedback_text, source)
        VALUES (?, ?, ?, ?)
    """, (video_id, result_id, feedback_text, source))

    feedback_id = cur.lastrowid
    conn.commit()
    conn.close()

    print(f"[DB] 피드백 저장 완료 (feedback_id={feedback_id})")
    return feedback_id


# ════════════════════════════════════════════════════════════
# 3. 관리자 전용 내보내기
# ════════════════════════════════════════════════════════════
def _fetch_all(date_from: str = None, date_to: str = None):
    """
    내부 공통 조회 함수.
    date_from / date_to : "YYYY-MM-DD" 형식, None 이면 전체 조회
    """
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur  = conn.cursor()

    query  = "SELECT * FROM user_feedback WHERE 1=1"
    params = []

    if date_from:
        query  += " AND created_at >= ?"
        params.append(f"{date_from} 00:00:00")
    if date_to:
        query  += " AND created_at <= ?"
        params.append(f"{date_to} 23:59:59")

    query += " ORDER BY created_at DESC"
    cur.execute(query, params)

    rows = cur.fetchall()
    conn.close()
    return rows


# ── 3-1. CSV 내보내기 ────────────────────────────────────────
def export_csv(date_from: str = None,
               date_to: str   = None,
               filename: str  = None) -> str:
    """
    CSV 파일로 내보냅니다.
    Returns : 저장된 파일 경로
    """
    rows = _fetch_all(date_from, date_to)
    if not rows:
        print("[EXPORT] 조회된 데이터가 없습니다.")
        return None

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"user_feedback_{timestamp}.csv"

    filepath = os.path.join(EXPORT_DIR, filename)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows([dict(row) for row in rows])

    print(f"[EXPORT] CSV 저장 완료 → {filepath}  ({len(rows)}건)")
    return filepath


# ── 3-2. JSON 내보내기 ───────────────────────────────────────
def export_json(date_from: str = None,
                date_to: str   = None,
                filename: str  = None) -> str:
    """
    JSON 파일로 내보냅니다.
    Returns : 저장된 파일 경로
    """
    rows = _fetch_all(date_from, date_to)
    if not rows:
        print("[EXPORT] 조회된 데이터가 없습니다.")
        return None

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"user_feedback_{timestamp}.json"

    filepath = os.path.join(EXPORT_DIR, filename)
    data     = [dict(row) for row in rows]

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"[EXPORT] JSON 저장 완료 → {filepath}  ({len(rows)}건)")
    return filepath


# ── 3-3. Excel 내보내기 (openpyxl 필요) ─────────────────────
def export_excel(date_from: str = None,
                 date_to: str   = None,
                 filename: str  = None) -> str:
    """
    Excel(.xlsx) 파일로 내보냅니다.
    openpyxl 이 설치되어 있어야 합니다: pip install openpyxl
    Returns : 저장된 파일 경로
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[EXPORT] openpyxl 이 설치되지 않았습니다. pip install openpyxl")
        return None

    rows = _fetch_all(date_from, date_to)
    if not rows:
        print("[EXPORT] 조회된 데이터가 없습니다.")
        return None

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"user_feedback_{timestamp}.xlsx"

    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "User Feedback"

    # 헤더 스타일
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(fill_type="solid", fgColor="2E4A8B")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = list(rows[0].keys())
    col_widths = {
        "feedback_id":   12,
        "video_id":      28,
        "result_id":     28,
        "feedback_text": 60,
        "source":        20,
        "created_at":    22,
    }

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = col_widths.get(header, 15)

    ws.row_dimensions[1].height = 22

    # 데이터 행
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(dict(row).values(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    print(f"[EXPORT] Excel 저장 완료 → {filepath}  ({len(rows)}건)")
    return filepath


# ════════════════════════════════════════════════════════════
# 4. 간단 조회 유틸
# ════════════════════════════════════════════════════════════
def get_summary():
    """전체 피드백 요약 통계를 출력합니다."""
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM user_feedback")
    total = cur.fetchone()[0]

    cur.execute("""
        SELECT source, COUNT(*) as cnt
        FROM user_feedback
        GROUP BY source
        ORDER BY cnt DESC
    """)
    by_source = cur.fetchall()

    cur.execute("""
        SELECT DATE(created_at) as day, COUNT(*) as cnt
        FROM user_feedback
        GROUP BY day
        ORDER BY day DESC
        LIMIT 7
    """)
    by_day = cur.fetchall()

    conn.close()

    print("=" * 40)
    print(f"  User Feedback 요약")
    print("=" * 40)
    print(f"  전체 피드백 수 : {total}건")
    print()
    print("  [유입 경로별]")
    for source, cnt in by_source:
        print(f"    {source:<25} {cnt}건")
    print()
    print("  [최근 7일 일별]")
    for day, cnt in by_day:
        print(f"    {day}  {cnt}건")
    print("=" * 40)


# ════════════════════════════════════════════════════════════
# 5. CLI 실행 예시
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys

    # DB 초기화 (최초 1회)
    init_db()

    # 테스트 데이터 삽입
    if "--test" in sys.argv:
        save_feedback(
            video_id="dQw4w9WgXcQ",
            feedback_text="이 영상은 정상인데 부적합으로 판정됐습니다. 재검토 부탁드립니다.",
            result_id="RES_001"
        )
        save_feedback(
            video_id="abc123xyz",
            feedback_text="낚시성 제목인데 정상으로 분류된 것 같습니다.",
            result_id="RES_002"
        )
        print("[TEST] 테스트 데이터 2건 삽입 완료\n")

    # 요약 출력
    get_summary()

    print()
    print("내보내기 실행 중...")

    # 전체 기간 CSV / JSON / Excel 동시 내보내기
    export_csv()
    export_json()
    export_excel()

    print()
    print("특정 기간 필터 예시 (2026-03-01 ~ 2026-03-09):")
    export_csv(date_from="2026-03-01", date_to="2026-03-09",
               filename="feedback_march.csv")
