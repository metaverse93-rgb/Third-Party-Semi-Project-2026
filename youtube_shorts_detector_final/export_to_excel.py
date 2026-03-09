"""
DB 분석 결과를 엑셀로 내보내기
실행: python export_to_excel.py
"""
from database_manager import db_manager
from database_models import AnalysisResults, Contents
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

CATEGORY_COLORS = {
    "C1": "FFCCCC",
    "C2": "FFE5CC",
    "C3": "FFFFCC",
    "C4": "E5CCFF",
    "C5": "CCFFCC",
}

CATEGORY_NAMES = {
    "C1": "어그로/스팸",
    "C2": "공장형 패턴",
    "C3": "품질 불량",
    "C4": "무단 도용",
    "C5": "정상 영상",
}

def export():
    rows = []
    stats = {}

    with db_manager.get_db_session() as session:
        results = session.query(AnalysisResults).order_by(AnalysisResults.created_at.desc()).all()
        contents = {c.video_id: c for c in session.query(Contents).all()}

        for r in results:
            content = contents.get(r.video_id)
            cat = r.c_category
            if cat not in stats:
                stats[cat] = {"count": 0, "confidence_sum": 0}
            stats[cat]["count"] += 1
            stats[cat]["confidence_sum"] += r.confidence_score

            # performance_metrics JSON에서 점수 추출
            pm = r.performance_metrics or {}
            c1_score  = pm.get("c1_spam_score", "")
            c2_score  = pm.get("c2_pattern_score", "")
            c3_score  = pm.get("c3_context_score", "")
            cis_score = pm.get("cis_final", "")

            rows.append({
                "video_id":       r.video_id,
                "url":            content.url if content else f"https://youtube.com/shorts/{r.video_id}",
                "title":          content.title if content else "",
                "channel_name":   content.channel_name if content else "",
                "c_category":     cat,
                "category_name":  CATEGORY_NAMES.get(cat, ""),
                "confidence_score": r.confidence_score,
                "c1_score":       c1_score,
                "c2_score":       c2_score,
                "c3_score":       c3_score,
                "cis_score":      cis_score,
                "status":         r.status.value if r.status else "",
                "processing_time": round(r.processing_time, 2) if r.processing_time else "",
                "created_at":     r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
                "reasoning_log":  (r.reasoning_log or "")[:200],
            })

    if not rows:
        print("❌ 저장된 분석 결과가 없습니다.")
        return

    wb = Workbook()

    # =============================================
    # 시트 1: 전체 분석 결과
    # =============================================
    ws1 = wb.active
    ws1.title = "분석 결과"

    headers = [
        "No", "Video ID", "URL", "제목", "채널명",
        "카테고리", "카테고리명", "신뢰도",
        "C1 스팸점수", "C2 공장형점수", "C3 맥락점수", "CIS 최종점수",
        "상태", "처리시간(초)", "분석일시", "판단 근거"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="2D5A8E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    for row_idx, r in enumerate(rows, 2):
        color = CATEGORY_COLORS.get(r["c_category"], "FFFFFF")
        row_data = [
            row_idx - 1,
            r["video_id"],
            r["url"],
            r["title"],
            r["channel_name"],
            r["c_category"],
            r["category_name"],
            r["confidence_score"],
            r["c1_score"],
            r["c2_score"],
            r["c3_score"],
            r["cis_score"],
            r["status"],
            r["processing_time"],
            r["created_at"],
            r["reasoning_log"],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 15))
            # 카테고리 컬럼 색상
            if col_idx in [5, 6]:
                cell.fill = PatternFill("solid", start_color=color)
            # 점수 컬럼 소수점 포맷
            if col_idx in [7, 8, 9, 10, 11]:
                cell.number_format = "0.000"
            # CIS 점수는 음수 가능하므로 별도 포맷
            if col_idx == 11 and isinstance(val, (int, float)):
                cell.number_format = "0.000;-0.000"
        ws1.row_dimensions[row_idx].height = 18

    col_widths = [6, 15, 40, 35, 20, 10, 14, 10, 13, 13, 13, 13, 14, 12, 20, 60]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    # =============================================
    # 시트 2: 카테고리별 통계
    # =============================================
    ws2 = wb.create_sheet("통계 요약")

    stat_headers = ["카테고리", "카테고리명", "분석 건수", "비율", "평균 신뢰도"]
    for col, h in enumerate(stat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="2D5A8E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    total = len(rows)
    for row_idx, cat in enumerate(["C1", "C2", "C3", "C4", "C5"], 2):
        s = stats.get(cat, {"count": 0, "confidence_sum": 0})
        count = s["count"]
        avg_conf = s["confidence_sum"] / count if count > 0 else 0
        color = CATEGORY_COLORS.get(cat, "FFFFFF")

        row_data = [cat, CATEGORY_NAMES[cat], count, count / total if total > 0 else 0, avg_conf]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", start_color=color)
            if col_idx == 4:
                cell.number_format = "0.0%"
            if col_idx == 5:
                cell.number_format = "0.000"
        ws2.row_dimensions[row_idx].height = 20

    ws2.cell(row=7, column=1, value="합계").font = Font(bold=True, name="Arial")
    ws2.cell(row=7, column=3, value="=SUM(C2:C6)").font = Font(bold=True, name="Arial")
    cell_total = ws2.cell(row=7, column=4, value="=SUM(D2:D6)")
    cell_total.font = Font(bold=True, name="Arial")
    cell_total.number_format = "0.0%"

    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # =============================================
    # 저장
    # =============================================
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"shorts_analysis_{timestamp}.xlsx"
    wb.save(filename)
    print(f"✅ 엑셀 저장 완료: {filename}")
    print(f"   총 {len(rows)}개 | 카테고리 분포: { {k: v['count'] for k, v in stats.items()} }")

if __name__ == "__main__":
    export()
